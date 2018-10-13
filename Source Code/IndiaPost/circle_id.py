import csv
from datetime import datetime, timedelta
import openpyxl
import time
import math
import os
import pandas as pd
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles import PatternFill
import xlsxwriter

curDir = "C:\\web2py_win\\web2py\\applications\\IndiaPost_UI\\input files\\"
pwd = "C:\\web2py_win\\web2py\\applications\\IndiaPost_UI\\Source Code\\IndiaPost\\"
df_loc = pd.read_excel(curDir+'Location_Details.xlsx') #Reading the Values of 369 locations
loc_id_list = df_loc.values.tolist() #Creating a list loc_id_list for 369 locations
num_loc = len(loc_id_list)
print(num_loc)

fp_out = openpyxl.load_workbook(pwd + "Circle_ID.xlsx")

sheet_out = fp_out.active

sheet_out.title =  "Circle_ID"

sheet_out.cell(row = 1, column = 1).value = "Circle Name"
sheet_out.cell(row = 1, column = 2).value = "ID"
sheet_out.cell(row = 2, column = 1).value = "National"
sheet_out.cell(row = 2, column = 2).value = 0
sheet_out.cell(row = 3, column = 1).value = "Andhra Pradesh"
sheet_out.cell(row = 3, column = 2).value = 1
sheet_out.cell(row = 4, column = 1).value = "Assam"
sheet_out.cell(row = 4, column = 2).value = 2
sheet_out.cell(row = 5, column = 1).value = "Bihar"
sheet_out.cell(row = 5, column = 2).value = 3
sheet_out.cell(row = 6, column = 1).value = "Chhattisgarh"
sheet_out.cell(row = 6, column = 2).value = 4
sheet_out.cell(row = 7, column = 1).value = "Delhi"
sheet_out.cell(row = 7, column = 2).value = 5
sheet_out.cell(row = 8, column = 1).value = "Goa"
sheet_out.cell(row = 8, column = 2).value = 6
sheet_out.cell(row = 9, column = 1).value = "Gujarat"
sheet_out.cell(row = 9, column = 2).value = 7
sheet_out.cell(row = 10, column = 1).value = "Haryana"
sheet_out.cell(row = 10, column = 2).value = 8
sheet_out.cell(row = 11, column = 1).value = "Himachal Pradesh"
sheet_out.cell(row = 11, column = 2).value = 9
sheet_out.cell(row = 12, column = 1).value = "Jammu & Kashmir"
sheet_out.cell(row = 12, column = 2).value = 10
sheet_out.cell(row = 13, column = 1).value = "Jharkhand"
sheet_out.cell(row = 13, column = 2).value = 11
sheet_out.cell(row = 14, column = 1).value = "Karnataka"
sheet_out.cell(row = 14, column = 2).value = 12
sheet_out.cell(row = 15, column = 1).value = "Kerala"
sheet_out.cell(row = 15, column = 2).value = 13
sheet_out.cell(row = 16, column = 1).value = "Madhya Pradesh"
sheet_out.cell(row = 16, column = 2).value = 14
sheet_out.cell(row = 17, column = 1).value = "Maharashtra"
sheet_out.cell(row = 17, column = 2).value = 15
sheet_out.cell(row = 18, column = 1).value = "North East"
sheet_out.cell(row = 18, column = 2).value = 16
sheet_out.cell(row = 19, column = 1).value = "Orissa"
sheet_out.cell(row = 19, column = 2).value = 17
sheet_out.cell(row = 20, column = 1).value = "Punjab"
sheet_out.cell(row = 20, column = 2).value = 18
sheet_out.cell(row = 21, column = 1).value = "Rajasthan"
sheet_out.cell(row = 21, column = 2).value = 19
sheet_out.cell(row = 22, column = 1).value = "Tamil Nadu"
sheet_out.cell(row = 22, column = 2).value = 20
sheet_out.cell(row = 23, column = 1).value = "Uttar Pradesh"
sheet_out.cell(row = 23, column = 2).value = 21
sheet_out.cell(row = 24, column = 1).value = "Uttarakhand"
sheet_out.cell(row = 24, column = 2).value = 22
sheet_out.cell(row = 25, column = 1).value = "West Bengal"
sheet_out.cell(row = 25, column = 2).value = 23

fp_out.save("C:\\web2py_win\\web2py\\applications\\IndiaPost_UI\\Source Code\\IndiaPost\\Circle_ID.xlsx")

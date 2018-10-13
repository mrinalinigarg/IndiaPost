import csv
from datetime import datetime, timedelta
import openpyxl
import time
import math
import os
import pandas as pd
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles import PatternFill

# Declaring global scope of commonly used lists and dicts
curDir = "C:\\web2py_win\\web2py\\applications\\IndiaPost_UI\\Source Code\\IndiaPost\\"
with open(curDir+'loc_to_id.csv', 'r') as fp_in:
	reader = csv.reader(fp_in)
	list_loc_to_id = list(reader)

num_locations = len(list_loc_to_id)
#print(num_locations)

dict_loc_to_id = {}
for i in range(len(list_loc_to_id)):
	dict_loc_to_id[list_loc_to_id[i][0]] = int(list_loc_to_id[i][1])

df_ln = pd.read_excel(curDir+'Location_names.xlsx')#Reading corrected names of airport locations
list_ln = df_ln.values.tolist()
dict_ln = {}#Creating a Dictionary to assign corrected names to the airport locations
for i in range(len(list_ln)):
	dict_ln[list_ln[i][0]] = list_ln[i][1]

fp_tp = openpyxl.load_workbook(curDir+'TP_Format.xlsx')#Reading Flight Schedule for all flights
sheet_tp = fp_tp["Sheet1"]
num_rows_tp = sheet_tp.max_row 

with open(curDir+'air.csv', 'r') as fp_in:
	reader = csv.reader(fp_in)
	list_air = list(reader)

with open(curDir+'output_train.csv', 'r') as fp_in:
	reader = csv.reader(fp_in)
	list_rail = list(reader)

with open(curDir+'surf_matrix.csv', 'r') as fp_in:
	reader = csv.reader(fp_in)
	list_surf_distance = list(reader)

with open(curDir+'loc_to_circle.csv', 'r') as fp_in:
	reader = csv.reader(fp_in)
	list_loc_to_circle = list(reader)

with open(curDir+'circle_to_transit.csv', 'r') as fp_in:
	reader = csv.reader(fp_in)
	list_circle_to_transit = list(reader)

with open(curDir+'closest_airport.csv', 'r') as fp_in:
	reader = csv.reader(fp_in)
	list_closest_air = list(reader)

df_loc = pd.read_excel(curDir+'SPH-NPH Mapping.xlsx') #Reading the SPH-NPH Mapping
loc_id_list = df_loc.values.tolist() 
num_loc = len(loc_id_list) 
	
dict_id = {} #Creating a Dictionary for SPH-NPH Mapping
for i in range(num_loc):
	dict_id[loc_id_list[i][1]] = loc_id_list[i][2]

'''
buffer_origin_air = 3; buffer_origin_surf = 1
buffer_dest_air = 3; buffer_dest_surf = 1
buffer_tp_air = 2; buffer_tp_surf = 2
buffer_nph_dep =4 ; buffer_nph_arr =4
surf_speed = 30


'''


buff_in = openpyxl.load_workbook(curDir +'Buffer.xlsx') # buffer input automation
sheet_in_buff = buff_in.get_sheet_by_name('Sheet1')


buffer_origin_air = int(sheet_in_buff.cell(row = 2, column = 1).value)
buffer_origin_surf = int(sheet_in_buff.cell(row = 2, column = 2).value)
buffer_dest_air = int(sheet_in_buff.cell(row = 2, column = 3).value)
buffer_dest_surf = int(sheet_in_buff.cell(row = 2, column = 4).value)
buffer_tp_air = int(sheet_in_buff.cell(row = 2, column = 5).value)
buffer_tp_surf = int(sheet_in_buff.cell(row = 2, column = 6).value)
buffer_nph_dep = int(sheet_in_buff.cell(row = 2, column = 7).value)
buffer_nph_arr = int(sheet_in_buff.cell(row = 2, column = 8).value)
surf_speed = int(sheet_in_buff.cell(row = 2, column = 9).value)

def surface_tp(source_id,des_id):
	
	dict_err_src = {} #Initialising source Dictionary
	dict_err_dest = {} #Initialising destination Dictionary

	for i in range(num_rows_tp):
		name_src = str(sheet_tp.cell(row = i+1, column = 1).value)
		name_dest = str(sheet_tp.cell(row = i+1, column = 2).value)
		number_of_rows = int(sheet_tp.cell(row = i+1, column = 3).value)
		#print(name_src,name_dest,number_of_rows)
		curr_col = 4
		src_id = dict_loc_to_id[dict_ln[name_src]] # Unique ID of Source Location
		dest_id = dict_loc_to_id[dict_ln[name_dest]]
		
		if(src_id==source_id) and (dest_id==des_id):
			TP = [[] for x in range(number_of_rows)]
			for j in range(number_of_rows):
				number_of_tps = int(sheet_tp.cell(row = i+1, column = curr_col).value)
				curr_col+=1

				TP_list = [[] for x in range(number_of_rows)]

				for k in range(number_of_tps):

					tp = int(sheet_tp.cell(row = i+1, column = curr_col).value)
					curr_col+=1
					if(tp!=src_id) and (tp!=dest_id):
						TP_list[j].append(tp)
			
				TP[j].extend(TP_list[j])
				#print(TP_list)
				
			#print(name_src,name_dest,number_of_rows)
			#print(TP)
			
			return(TP,number_of_rows)

def closest_arrival(dept_datetime, list_mode, buff_dept, buff_arrival):
	

	list_timestamps = []

	total_index = int(list_mode[2]) # no. of row entries
	if(total_index == 0):
		inf_datetime = datetime(9999, 12, 31, 23, 59, 59)
		return(inf_datetime, inf_datetime, inf_datetime)# if no connections then return infinite time

	FMT = '%H%M'
	curr_datetime = dept_datetime + timedelta(hours = buff_dept) # current time with buffer for departure
	curr_day = curr_datetime.day # datetime weekday
	curr_weekday = curr_datetime.weekday()
	curr_weekday_bin = 2**(curr_weekday) # weekday function starts with monday = 0 ... sunday = 6
	curr_time = curr_datetime.strftime(FMT)#converting time format

	curr_index = 3
	min_transit_datetime = datetime(9999, 12, 31, 23, 59, 59)#intializing infinite value to compare for minimum
	for mode_id in range(total_index):
		dept_schedule = int(list_mode[curr_index])
		dept_time = str(list_mode[curr_index+1])
		transit_duration = float(list_mode[curr_index+3])
		dept_day = curr_day
		dept_weekday = curr_weekday
	
		if(((dept_schedule & curr_weekday_bin) > 0) and (curr_time <= dept_time)): #if the transit is valid for that day and current time is less than departure time we will replace the values as minimun transit time
			transit_datetime = datetime(1, 1, dept_day, int(dept_time[:2]), int(dept_time[2:])) + timedelta(hours = transit_duration)
		else:
			while(True):
				dept_weekday += 1 # check for next day connections for the entire week
				dept_day += 1
				dept_weekday_bin = 2**(dept_weekday % 7)
				if((dept_schedule & dept_weekday_bin) > 0):#connection found for transit
					transit_datetime = datetime(1, 1, dept_day, int(dept_time[:2]), int(dept_time[2:])) + timedelta(hours = transit_duration)
					break
		if(transit_datetime < min_transit_datetime):
			min_transit_datetime = transit_datetime
			min_transit_hours = transit_duration # journey time corresponding to shortest arrival
		curr_index += 4

	dept_datetime_mode = min_transit_datetime - timedelta(hours = min_transit_hours) # departure time
	arr_datetime_wb = min_transit_datetime + timedelta(hours = buff_arrival) # adding arrival buffer

	return(arr_datetime_wb, dept_datetime_mode, min_transit_datetime)

def dept_road(dept_datetime, surf_distance, buff_dept, buff_arrival):

	dept_datetime_mode = dept_datetime + timedelta(hours = buff_dept)
	min_transit_datetime = dept_datetime_mode + timedelta(hours = (surf_distance/surf_speed))
	arr_datetime_wb = min_transit_datetime + timedelta(hours = buff_arrival)

	return(arr_datetime_wb, dept_datetime_mode, min_transit_datetime)


def opt_transit_nph_all(source_id, dest_id, dispatch_datetime,buff_dept, buff_arrival):


	list_transit_timestamps = []
	list_transit_locations = []
	list_transit_modes = []	
	arrival_datetime = datetime(9999, 12, 31, 23, 59)
	#source_id_proxy = int(list_closest_air[source_id][0]); dest_id_proxy = int(list_closest_air[dest_id][0]) # Mapped IDs according to nearest airport
	#source_circle_id = int(list_loc_to_circle[source_id_proxy][0]); dest_circle_id = int(list_loc_to_circle[dest_id_proxy][0])

	list_transit_all,no_tp_rows =surface_tp(source_id,dest_id)
	#print(list_transit_all)
	#print(no_tp_rows)


	#source_circle_id = int(list_loc_to_circle[source_id_proxy][0]); dest_circle_id = int(list_loc_to_circle[dest_id_proxy][0])
	
	LTL_TP = []
	LTM_TP = []
	LTT_TP = []
	AT_TP = []

	if(no_tp_rows == 0):
		
		#return(opt_transit_nph_new(source_id, dest_id, dispatch_datetime))
		datetime_arr_wb, datetime_dept, datetime_arr = dept_road(dispatch_datetime, float(list_surf_distance[source_id][dest_id]), buff_dept, buff_arrival)
		list_transit_timestamps.extend([datetime_dept, datetime_arr, datetime_arr_wb])
		list_transit_modes.append('0')
				
		return(list_transit_locations, list_transit_modes, list_transit_timestamps, datetime_arr_wb)
		#list_transit_locations_new, list_transit_modes_new, list_transit_timestamps_new, arrival_datetime_new = opt_transit_nph_new(source_id, dest_id, dispatch_datetime)
	
	elif(no_tp_rows > 0):
		datetime_opt_arrival = datetime(9999, 12, 31, 23, 59)
		list_opt_timestamps = []
		list_opt_loc = []
		list_opt_mode = []

		for i in range(no_tp_rows):
			list_transit=list_transit_all[i]
			length_transit = len(list_transit)
			
			print(list_transit,len(list_transit))

			datetime_opt_arrival_tp = datetime(9999, 12, 31, 23, 59)
			list_opt_timestamps_tp = []
			list_opt_loc_tp = []
			list_opt_mode_tp = []

			if(source_id == dest_id):
				is_valid_transit = False
			else:
				is_valid_transit = True

			if(is_valid_transit is False):

				datetime_arr_wb, datetime_dept, datetime_arr = dept_road(dispatch_datetime, float(list_surf_distance[source_id][dest_id]), buff_dept, buff_arrival)
				list_transit_timestamps.extend([datetime_dept, datetime_arr, datetime_arr_wb])
				list_transit_modes.append('0')
				
				return(list_transit_locations, list_transit_modes, list_transit_timestamps, datetime_arr_wb)

			list_timestamps_tp = []
			
			list_mode_tp = []	

			index_leg = source_id*num_locations + list_transit[0]
			print(index_leg,source_id,list_transit[0])
			datetime_arr_road_wb, datetime_dept_road, datetime_arr_road = dept_road(dispatch_datetime, float(list_surf_distance[source_id][list_transit[0]]), buff_dept, buffer_tp_surf)
			#datetime_arr_air_wb, datetime_dept_air, datetime_arr_air = closest_arrival(dispatch_datetime, list_air[index_leg], buffer_origin_air_pr, buffer_tp_air)
			datetime_arr_rail_wb, datetime_dept_rail, datetime_arr_rail = closest_arrival(dispatch_datetime, list_rail[index_leg], buff_dept, buffer_tp_surf)					
			list_mode_1 = min([[datetime_arr_road_wb, datetime_dept_road, datetime_arr_road, '0'], [datetime_arr_rail_wb, datetime_dept_rail, datetime_arr_rail, '2']])					
			datetime_arrival = list_mode_1[0]
			list_timestamps_tp.extend([list_mode_1[1], list_mode_1[2], list_mode_1[0]])
			list_mode_tp.append(list_mode_1[3])

			if(len(list_transit)>1):
				

				for j in range(length_transit-1):
					
					index_leg = list_transit[j]*num_locations + list_transit[j+1]
					print(index_leg,list_transit[j],list_transit[j+1])
					datetime_arr_road_wb, datetime_dept_road, datetime_arr_road = dept_road(datetime_arrival, float(list_surf_distance[list_transit[j]][list_transit[j+1]]), buffer_tp_surf, buffer_tp_surf)
					#datetime_arr_air_wb, datetime_dept_air, datetime_arr_air = closest_arrival(dispatch_datetime, list_air[index_leg], buffer_origin_air_pr, buffer_tp_air)
					datetime_arr_rail_wb, datetime_dept_rail, datetime_arr_rail = closest_arrival(datetime_arrival, list_rail[index_leg], buffer_tp_surf, buffer_tp_surf)					
					list_mode_2 = min([[datetime_arr_road_wb, datetime_dept_road, datetime_arr_road, '0'], [datetime_arr_rail_wb, datetime_dept_rail, datetime_arr_rail, '2']])					
					#list_mode = min([[list_mode_2[0],list_mode_2[1], list_mode_2[2],list_mode_2[3]],[list_mode[0],list_mode[1], list_mode[2],list_mode[3]]])
					datetime_arrival = list_mode_2[0]
					list_timestamps_tp.extend([list_mode_2[1], list_mode_2[2], list_mode_2[0]])
					list_mode_tp.append(list_mode_2[3])

			
			index_leg = list_transit[length_transit-1]*num_locations + dest_id
			print(index_leg,list_transit[length_transit-1],dest_id)
			datetime_arr_road_wb, datetime_dept_road, datetime_arr_road = dept_road(datetime_arrival, float(list_surf_distance[list_transit[length_transit-1]][dest_id]), buffer_tp_surf, buff_arrival)
			#datetime_arr_air_wb, datetime_dept_air, datetime_arr_air = closest_arrival(dispatch_datetime, list_air[index_leg], buffer_origin_air_pr, buffer_tp_air)
			datetime_arr_rail_wb, datetime_dept_rail, datetime_arr_rail = closest_arrival(datetime_arrival, list_rail[index_leg], buffer_tp_surf, buff_arrival)					
			list_mode_2 = min([[datetime_arr_road_wb, datetime_dept_road, datetime_arr_road, '0'], [datetime_arr_rail_wb, datetime_dept_rail, datetime_arr_rail, '2']])					
			datetime_arrival = list_mode_2[0]
			list_timestamps_tp.extend([list_mode_2[1], list_mode_2[2], list_mode_2[0]])
			list_mode_tp.append(list_mode_2[3])	

			if(datetime_arrival < datetime_opt_arrival_tp):
				datetime_opt_arrival_tp = datetime_arrival
				list_opt_timestamps_tp = list_timestamps_tp
				list_opt_loc_tp = list_transit
				list_opt_mode_tp = list_mode_tp
			
			print(datetime_opt_arrival_tp,list_opt_timestamps_tp,list_opt_loc_tp,list_opt_mode_tp)

			if(datetime_opt_arrival_tp<datetime_opt_arrival):
				datetime_opt_arrival = datetime_opt_arrival_tp
				list_opt_timestamps = list_opt_timestamps_tp
				list_opt_loc = list_opt_loc_tp
				list_opt_mode = list_opt_mode_tp
		print(datetime_opt_arrival,list_opt_timestamps,list_opt_loc,list_opt_mode)

		return(list_opt_loc, list_opt_mode, list_opt_timestamps,datetime_opt_arrival)

def main():



	fp_in = openpyxl.load_workbook(curDir +'NPH to SPH.xlsx')
	sheet_in = fp_in.get_sheet_by_name('Sheet')
	num_rows = sheet_in.max_row - 1

	if(num_rows>0):

		list_travel_mode = ['Road', 'Air', 'Rail']
		list_days_week = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
		weekday_dict={'Mon':1, 'Tue':2, 'Wed':3, 'Thu':4, 'Fri':5, 'Sat':6, 'Sun':7} 

		fp_out = openpyxl.Workbook()
		sheet_out = fp_out.active

		num_timestamps = 0

		num_tp = 0
		#sheet_out.append(['Origin','Destination','Dispatch Time','Dispatch Day','Dept. from Origin','Day of Dept.','Arrival at TP1','Day of Arrival','Arrival at TP1 with buffer','Day of Arrival','Mode of Travel','TP1','Dept. from TP1','Day of Dept','Arrival at TP2','Day of Arrival','Arrival at TP2 With buffer','Day of Arrival','Mode of Travel','TP2','Dept. from TP2','Day of Dept','Arrival at TP3','Day of Arrival','Arrival at TP3 With buffer','Day of Arrival','Mode of Travel','TP3','Dept. from TP3','Day of Dept','Arrival at TP4','Day of Arrival','Arrival at TP4 With buffer','Day of Arrival','Mode of Travel','TP4','Dept. from TP4','Day of Dept','Arrival at TP5','Day of Arrival','Arrival at TP5 With buffer','Day of Arrival','Mode of Travel','TP5','Dept. from TP5','Day of Dept','Arrival at TP6','Day of Arrival','Arrival at TP6 With buffer','Day of Arrival','Mode of Travel','Destination','TotalTransit Time', 'Destination NPH'])
		sheet_out.append(['Serial Number','Origin','Destination','Dispatch Time','Dispatch Day','Destination NPH','TotalTransit Time','Dept. from Origin','Day of Dept.','Arrival at TP1','Day of Arrival','Arrival at TP1 with buffer','Day of Arrival','Mode of Travel','TP1','Dept. from TP1','Day of Dept','Arrival at TP2','Day of Arrival','Arrival at TP2 With buffer','Day of Arrival','Mode of Travel','TP2','Dept. from TP2','Day of Dept','Arrival at TP3','Day of Arrival','Arrival at TP3 With buffer','Day of Arrival','Mode of Travel','TP3','Dept. from TP3','Day of Dept','Arrival at TP4','Day of Arrival','Arrival at TP4 With buffer','Day of Arrival','Mode of Travel','TP4','Dept. from TP4','Day of Dept','Arrival at TP5','Day of Arrival','Arrival at TP5 With buffer','Day of Arrival','Mode of Travel','TP5','Dept. from TP5','Day of Dept','Arrival at TP6','Day of Arrival','Arrival at TP6 With buffer','Day of Arrival','Mode of Travel','Destination'])
		
		for i in range(num_rows):
			serial_number = str(sheet_in.cell(row = i+2, column = 1).value)
			source_name_NPH = str(sheet_in.cell(row = i+2, column = 2).value)
			destination_name_SPH = str(sheet_in.cell(row = i+2, column = 3).value)
			#proxy_source_name_NPH = dict_id[source_name_SPH] 
			proxy_destination_name_NPH = dict_id[destination_name_SPH]

			source_NPH_id = dict_loc_to_id[source_name_NPH]
			dest_SPH_id = dict_loc_to_id[destination_name_SPH]
			#source_NPH_id = dict_loc_to_id[proxy_source_name_NPH]
			dest_NPH_id = dict_loc_to_id[proxy_destination_name_NPH]
			#source_name = str(sheet_in.cell(row = i+2, column = 2).value)
			#dest_name = str(sheet_in.cell(row = i+2, column = 3).value)
			#source_id = dict_loc_to_id[source_name]
			#dest_id = dict_loc_to_id[dest_name]
			
			dispatch_time = str(sheet_in.cell(row = i+2, column = 4).value)
			weekday_input = str(sheet_in.cell(row = i+2, column = 5).value)
			
			dispatch_first2dig=int(dispatch_time[0:2])
			dispatch_last2dig=int(dispatch_time[3:5])

			dispatch_datetime = datetime(1,1,weekday_dict[weekday_input],dispatch_first2dig,dispatch_last2dig)
			opt_overall=[]

			if(source_NPH_id ==dest_NPH_id):
				rail_arrival_datetime, rail_datetime_dept, rail_datetime_arr = closest_arrival(dispatch_datetime, list_rail[source_NPH_id*num_locations + dest_SPH_id], buffer_nph_dep, buffer_dest_surf)
				road_arrival_datetime, road_datetime_dept, road_datetime_arr = dept_road(dispatch_datetime, float(list_surf_distance[source_NPH_id][dest_SPH_id]), buffer_nph_dep, buffer_dest_surf)
				list_transit_locations, list_transit_modes, list_transit_timestamps, arrival_datetime = opt_transit_nph_all(source_NPH_id, dest_SPH_id, dispatch_datetime,buffer_nph_dep, buffer_dest_surf)

				opt_overall = min([arrival_datetime, list_transit_timestamps, list_transit_locations, list_transit_modes], [rail_arrival_datetime, [rail_datetime_dept, rail_datetime_arr, rail_arrival_datetime], [], ['2']], [road_arrival_datetime, [road_datetime_dept, road_datetime_arr, road_arrival_datetime], [], ['0']])

				#print("TRUE")
			else:

				list_transit_locations, list_transit_modes, list_transit_timestamps, arrival_datetime = opt_transit_nph_all(source_NPH_id, dest_NPH_id, dispatch_datetime,buffer_nph_dep, buffer_nph_arr)

				rail_arrival_datetime, rail_datetime_dept, rail_datetime_arr = closest_arrival(dispatch_datetime, list_rail[source_NPH_id*num_locations + dest_NPH_id], buffer_nph_dep, buffer_nph_arr)
				road_arrival_datetime, road_datetime_dept, road_datetime_arr = dept_road(dispatch_datetime, float(list_surf_distance[source_NPH_id][dest_NPH_id]), buffer_nph_dep, buffer_nph_arr)

				opt_overall_1 = min([arrival_datetime, list_transit_timestamps, list_transit_locations, list_transit_modes], [rail_arrival_datetime, [rail_datetime_dept, rail_datetime_arr, rail_arrival_datetime], [], ['2']], [road_arrival_datetime, [road_datetime_dept, road_datetime_arr, road_arrival_datetime], [], ['0']])

				rail_arrival_datetime, rail_datetime_dept, rail_datetime_arr = closest_arrival(opt_overall_1[0], list_rail[dest_NPH_id*num_locations + dest_SPH_id], buffer_nph_dep, buffer_dest_surf)
				road_arrival_datetime, road_datetime_dept, road_datetime_arr = dept_road(opt_overall_1[0], float(list_surf_distance[dest_NPH_id][dest_SPH_id]), buffer_nph_dep, buffer_dest_surf)
				list_transit_locations, list_transit_modes, list_transit_timestamps, arrival_datetime = opt_transit_nph_all(dest_NPH_id, dest_SPH_id, opt_overall_1[0],buffer_nph_dep, buffer_dest_surf)

				opt_overall_2 = min([arrival_datetime, list_transit_timestamps, list_transit_locations, list_transit_modes], [rail_arrival_datetime, [rail_datetime_dept, rail_datetime_arr, rail_arrival_datetime], [], ['2']], [road_arrival_datetime, [road_datetime_dept, road_datetime_arr, road_arrival_datetime], [], ['0']])

				
				opt_overall = opt_overall_1

				opt_overall[0] = opt_overall_2[0]
				opt_overall[1] = opt_overall_1[1]+opt_overall_2[1]
				opt_overall[2] = opt_overall_1[2]+ [dest_NPH_id] + opt_overall_2[2] 
				opt_overall[3] = opt_overall_1[3]+opt_overall_2[3]

				#print("FALSE")

			print(opt_overall)

					
			curr_col = 5
			
			isvalid_nts = True

			try:
				num_timestamps = len(opt_overall[1])
			except TypeError:
				isvalid_nts = False

			if(isvalid_nts):
				num_timestamps = len(opt_overall[1])
			else:
				num_timestamps = 0

			isvalid_ntp = True

			try:
				num_tp = len(opt_overall[2]) 
			except TypeError:
				isvalid_ntp = False

			if(isvalid_ntp):
				num_tp = len(opt_overall[2]) 
			else:
				num_tp = 0

			
			num_modes = num_tp + 1
			
		
			sheet_out.cell(row = i+2, column = 1).value = serial_number
			sheet_out.cell(row = i+2, column = 2).value = source_name_NPH
			sheet_out.cell(row = i+2, column = 3).value = destination_name_SPH
			sheet_out.cell(row = i+2, column = 4).value = dispatch_time
			sheet_out.cell(row = i+2, column = 5).value = weekday_input
			


			k = 0
			l = 0

			
			sheet_out.cell(row = i+2, column = 8).value = opt_overall[1][0].time()
			#print(source_name,dest_name ,opt_overall[1][0])
			curr_col += 1
			sheet_out.cell(row = i+2, column = 9).value = list_days_week[opt_overall[1][0].weekday()]#weekday
			curr_col += 1
			sheet_out.cell(row = i+2, column = 10).value = opt_overall[1][1].time()
			curr_col += 1
			sheet_out.cell(row = i+2, column = 11).value = list_days_week[opt_overall[1][1].weekday()]
			curr_col += 1
			sheet_out.cell(row = i+2, column = 12).value = opt_overall[1][2].time()
			curr_col += 1
			sheet_out.cell(row = i+2, column = 13).value = list_days_week[opt_overall[1][2].weekday()]
			curr_col += 1
			
			curr_col = 14
			
			curr_col_mode = curr_col
			curr_col_tran= curr_col + 1


			
			for j,k,l in zip(range(3,num_timestamps+2,3),range(num_tp),range(num_modes)):
				#print(num_timestamps)

				sheet_out.cell(row = i+2, column = curr_col_mode).value = list_travel_mode[int(opt_overall[3][l])]#all modes
				curr_col_mode += 8
				curr_col += 1
				l+=1
				sheet_out.cell(row = i+2, column = curr_col_tran).value = list_loc_to_id[int(opt_overall[2][k])][0]#transit point
				curr_col_tran += 8
				curr_col += 1
				k+=1
			
				sheet_out.cell(row = i+2, column = curr_col).value = opt_overall[1][j].time()
				curr_col += 1
				sheet_out.cell(row = i+2, column = curr_col).value = list_days_week[opt_overall[1][j].weekday()]#weekday
				curr_col += 1
				sheet_out.cell(row = i+2, column = curr_col).value = opt_overall[1][j+1].time()
				curr_col += 1
				sheet_out.cell(row = i+2, column = curr_col).value = list_days_week[opt_overall[1][j+1].weekday()]
				curr_col += 1
				sheet_out.cell(row = i+2, column = curr_col).value = opt_overall[1][j+2].time()
				curr_col += 1
				sheet_out.cell(row = i+2, column = curr_col).value = list_days_week[opt_overall[1][j+2].weekday()]
				curr_col += 1
				
				
			sheet_out.cell(row = i+2, column = curr_col).value = list_travel_mode[int(opt_overall[3][num_modes-1])]#all modes
			curr_col += 1
			sheet_out.cell(row = i+2, column = curr_col).value = destination_name_SPH
			curr_col += 1
			
			


			time_diff = timedelta.total_seconds((opt_overall[0]) - (dispatch_datetime))

			print(time_diff/3600)
			#sheet_out.cell(row = i+2, column = 70).value = (time_diff/3600)
			#sheet_out.cell(row = i+2, column = 71).value = source_name_NPH
			
			sheet_out.cell(row = i+2, column = 7).value = (time_diff/3600)
			sheet_out.cell(row = i+2, column = 7).fill = PatternFill(fgColor='FFEE08', fill_type = 'solid')
			sheet_out.cell(row = i+2, column = 6).value = proxy_destination_name_NPH 
			sheet_out.cell(row = i+2, column = 6).fill = PatternFill(fgColor='FFEE08', fill_type = 'solid')
			
			
			#sheet_out.cell(row = i+2, column = 72).value = proxy_destination_name_NPH 
			fp_out.save(curDir +'Surface_NPH to SPH.xlsx')
			#fp_out.save('C:\\web2py_win\\web2py\\applications\\IndiaPost\\private\\Output.xlsx')

			#fp_out.save('NPH_SPH_Surface_Output_12June7.xlsx')
			print(opt_overall[2], opt_overall[3], opt_overall[1], opt_overall[0])
	else:
		fp_out = openpyxl.Workbook()
		sheet_out = fp_out.active
		fp_out.save(curDir +'Surface_NPH to SPH.xlsx')
		fp_out.close()


	return

if __name__ == '__main__':
	main()
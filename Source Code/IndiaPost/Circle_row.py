import csv
from datetime import datetime, timedelta
import openpyxl 
import time
import math
import os
import pandas as pd
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles import PatternFill
import xlsxwriter
from openpyxl import Workbook

curDir = "C:\\web2py_win\\web2py\\applications\\IndiaPost_UI\\input files\\"
df_loc = pd.read_excel(curDir+'Location_Details.xlsx') #Reading the Values of 369 locations
loc_id_list = df_loc.values.tolist() #Creating a list loc_id_list for 369 locations
num_loc = len(loc_id_list)
print(num_loc)

#fp_out = openpyxl.load_workbook("Circle_Row.xlsx")

#sheet_out = fp_out.active

#sheet_out.title =  "Circle_Row"

data_list = [[] for x in range(num_loc*num_loc)]
data_list[0].append("Circle_ID")
data_list[1].append(0)
data_list[1].append('DELHI')
data_list[1].append('MUMBAI')
data_list[1].append('KOLKATA')
data_list[1].append('CHENNAI')
data_list[1].append('HYDERABAD')
data_list[1].append('BANGALORE')
data_list[1].append('GUWAHATI')
data_list[1].append('SILIGURI')
loc_circle=1

for i in range(1,24):
	#sheet_out.cell(row = i+3, column = 1).value =i+1
	data_list[i+1].append(i)

for i in range(num_loc):
	location_name = str(loc_id_list[i][1])
	location_id = int(loc_id_list[i][0])
	circle_name = str(loc_id_list[i][4])
	circle_id = int(loc_id_list[i][7])
	
	data_list[circle_id+1].append(location_name)
	
	#sheet_out.cell(row = circle_id+1, column = loc_circle+2).value = location_name
		
'''
with open('C:\\web2py_win\\web2py\\applications\\IndiaPost_UI\\Source Code\\IndiaPost\\coordinates.csv', 'w', newline = '') as fp_out: #Writing the values as list of lists in 1 go to decrease time complexity
	writer = csv.writer(fp_out)
	writer.writerows(data_list)



writer = pd.ExcelWriter('Circle_Row.xlsx', engine='xlsxwriter')
data_list.to_excel(writer, sheet_name='Sheet1')
writer.save()
'''
print(data_list)
#fp_out.save("Circle_Row.xlsx")

with open('C:\\web2py_win\\web2py\\applications\\IndiaPost_UI\\Source Code\\IndiaPost\\Circle_Row.csv', 'w', newline = '') as fp_out: #Writing the values as list of lists in 1 go to decrease time complexity
	writer = csv.writer(fp_out)
	writer.writerows(data_list)

wb = Workbook()
ws = wb.active
with open('C:\\web2py_win\\web2py\\applications\\IndiaPost_UI\\Source Code\\IndiaPost\\Circle_Row.csv', 'r') as f:
    for row in csv.reader(f):
        ws.append(row)
wb.save('C:\\web2py_win\\web2py\\applications\\IndiaPost_UI\\Source Code\\IndiaPost\\Circle_Row.xlsx')
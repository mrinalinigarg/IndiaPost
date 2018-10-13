import csv
from datetime import datetime, timedelta
import openpyxl
import time
import math
import os
import pandas as pd

def main():


	filename = "Surface.xlsx"

	curDir = "C:\\web2py_win\\web2py\\applications\\IndiaPost_UI\\input files\\"
	pwd = "C:\\web2py_win\\web2py\\applications\\IndiaPost_UI\\Source Code\\IndiaPost\\"
	df_loc = pd.read_excel(curDir+'Location_Details.xlsx') #Reading the Values of 369 locations
	loc_id_list = df_loc.values.tolist() #Creating a list loc_id_list for 369 locations
	num_loc = len(loc_id_list)
	#print(num_loc)

	#print(loc_id_list[0][1],loc_id_list[0][5])
	list_NPH=[]	
	list_SPH=[]

	for i in range(num_loc):
		if(loc_id_list[i][5]=="NPH"):
			list_NPH.append(loc_id_list[i][1])
		elif(loc_id_list[i][5]=="SPH"):
			list_SPH.append(loc_id_list[i][1])
	#print(list_NPH)
	#print(list_SPH)

	#fp_out = openpyxl.load_workbook(pwd+"Test.xlsx")

	#sheet_out = fp_out.active	
	#list_NPH = ['VIJAYAWADA','VISAKHAPATNAM','GUWAHATI','PATNA','MUZAFFARPUR','RAIPUR','DELHI','AHMEDABAD','SURAT','VADODARA','AMBALA','FARIDABAD','GURGAON','SHIMLA','PATHANKOT','JAMMU','SRINAGAR','JAMSHEDPUR','RANCHI','BANGALORE','HUBLI','MANGALORE','MYSORE','KOCHI','KOZHIKODE','TRIVANDRUM','BHOPAL','INDORE','AURANGABAD','MUMBAI','NAGPUR','PANAJI','PUNE','AGARTALA','AIZAWL','DIMAPUR','IMPHAL','BHUBANESWAR','CHANDIGARH','JALANDHAR','LUDHIANA','JAIPUR','CHENNAI','COIMBATORE','MADURAI','TRICHY','HYDERABAD','AGRA','ALLAHABAD','GHAZIABAD','LUCKNOW','BAREILLY','DEHRADUN','KOLKATA','SILIGURI','56 APO','99 APO']
	#list_SPH = ['TIRUPATHI','GUNTUR','NELLORE','KURNOOL','RAJAMUNDRY','ELURU','ANANTHAPUR','ONGOLE','CUDDAPAH','SRIKAKULAM','TEZPUR','SILCHAR','TINSUKIA','JORHAT','MUZAFFARPUR','BARAUNI','GAYA','BHAGALPUR','KIUL','CHAPRA','DURG','BILASPUR (CHHATTISGARH)','RAJKOT','VALSAD','BHARUCH','JAMNAGAR','JUNAGADH','ANAND','MEHASANA','BHAVNAGAR','BHUJ','HISAR','ROHTAK','KARNAL','SONEPAT','MANDI','HAMIRPUR','DHANBAD','BELGAUM','GULBARGA','DAVANAGERE','BIRUR','HASSAN','RAICHUR','TUMKUR','BAGALKOT','HOSPET','KOTTAYAM','THRISSUR','PALAKKAD','KOLLAM','TIRUVALLA','THODUPUZHA','KANNUR','JABALPUR','GWALIOR','SATNA','RATLAM','KATNI','ITARSI','MIRAJ','NASHIK','SOLAPUR','AHMEDNAGAR','JALGAON','PANVEL','SATARA','CHANDRAPUR','AKOLA','AMRAVATI','RATNAGIRI','NANDED','SHILLONG','ITANAGAR','CUTTACK','BERHAMPUR','SAMBALPUR','BALASORE','JEYPORE','BALANGIR','JAJPUR','PATIALA','AMRITSAR','BATHINDA','HOSHIARPUR','JODHPUR','KOTA','UDAIPUR','AJMER','BIKANER','SIKAR','FALNA','ALWAR','BHARATPUR','SRI GANGANAGAR','VELLORE','VRIDDHACHALAM','ERODE','THANJAVUR','VILLUPURAM','DINDIGUL','JOLARPETTAI','TIRUPUR','DHARMAPURI','TIRUNELVELI','MAYILADUTURAI','TUTICORIN','NAGARCOIL','KARAIKUDI','SALEM','WARANGAL','KARIMNAGAR','NALGONDA','KANPUR','ALIGARH','MEERUT','MORADABAD','GORAKHPUR','BAREILLY','SAHARANPUR','MAU','KHERI','VARANASI','RUDRAPUR','HARIDWAR','HOWRAH','SEALDAH','KHARAGPUR','BERHAMPORE(WB)','BURDWAN','MALDA','PORT BLAIR','ASANSOL']
	
	df_ln = pd.read_excel("C:\\web2py_win\\web2py\\applications\\IndiaPost_UI\\Source Code\\IndiaPost\\Location_names.xlsx")#Reading corrected names of airport locations
	
	list_ln = df_ln.values.tolist()
	dict_ln = {}#Creating a Dictionary to assign corrected names to the airport locations
	for i in range(len(list_ln)):
		dict_ln[list_ln[i][0]] = list_ln[i][1]



	'''
	fp_tp = openpyxl.load_workbook('SPH-NPH Mapping.xlsx')
	sheet_tp = fp_tp.get_sheet_by_name('SPH')
	num_rows_tp = sheet_tp.max_row -1
	print(sheet_tp[1][1])
	fp_tp2 = openpyxl.load_workbook('SPH-NPH Mapping.xlsx')
	sheet_tp2 = fp_tp2.get_sheet_by_name('NPH')
	num_rows_tp2 = sheet_tp2.max_row -1
	print(sheet_tp2[2])
	'''
	#list_NPH = ['VIJAYAWADA','VISAKHAPATNAM','GUWAHATI','PATNA','MUZAFFARPUR','RAIPUR','DELHI','AHMEDABAD','SURAT','VADODARA','AMBALA','FARIDABAD','GURGAON','SHIMLA','PATHANKOT','JAMMU','SRINAGAR','JAMSHEDPUR','RANCHI','BANGALORE','HUBLI','MANGALORE','MYSORE','KOCHI','KOZHIKODE','TRIVANDRUM','BHOPAL','INDORE','AURANGABAD','MUMBAI','NAGPUR','PANAJI','PUNE','AGARTALA','AIZAWL','DIMAPUR','IMPHAL','BHUBANESWAR','CHANDIGARH','JALANDHAR','LUDHIANA','JAIPUR','CHENNAI','COIMBATORE','MADURAI','TRICHY','HYDERABAD','AGRA','ALLAHABAD','GHAZIABAD','LUCKNOW','BAREILLY','DEHRADUN','KOLKATA','SILIGURI','56 APO','99 APO']
	#print(list_NPH[4])
	list_days_week = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']

	df_loc = pd.read_excel("C:\\web2py_win\\web2py\\applications\\IndiaPost_UI\\Source Code\\IndiaPost\\Surface.xlsx") 
	loc_id_list = df_loc.values.tolist() 
	num_loc = len(loc_id_list)

	df_loc2 = pd.read_excel("C:\\web2py_win\\web2py\\applications\\IndiaPost_UI\\Source Code\\IndiaPost\\location_validation.xlsx") 
	loc_id_list2 = df_loc2.values.tolist() 
	num_loc2 = len(loc_id_list2)
	#print(loc_id_list2)

	loc_id_list3 = [None] * (num_loc2)

	for i in range(num_loc2):
		loc_id_list3[i]=loc_id_list2[i][0]

	valid_counter_o = ""
	valid_counter_d = ""
	valid_counter_t = ""
	valid_counter_w = ""
	error = ""
	counter = 0
	ori=""
	dest=""

	for i in range(num_loc):
		row_num = loc_id_list[i][0]
		origin = loc_id_list[i][1]
		destination = loc_id_list[i][2]
		time = loc_id_list[i][3]
		week = loc_id_list[i][4]

		try:
			ori = dict_ln[origin]
			
		except KeyError:
			valid_counter_o = "Incorrect Origin Entry at : " +str(row_num)+ "."
			counter += 1
			error+=valid_counter_o
		try:
			
			dest = dict_ln[destination]
		except KeyError:
			valid_counter_d = "Incorrect Destination Entry at : " +str(row_num)+ "."
			counter += 1
			error+=valid_counter_d
		#print(origin,destination)
		try:
			time1 = int(time[0:2])
			time2 = str(time[2])
			time3 = int(time[3:5])
		except:
			print('An error occurred in Time Format.')

		if(ori in loc_id_list3):
			pass
		else:
			valid_counter_o = "Incorrect Origin Entry at : " +str(row_num)+ "."
			counter += 1
			error+=valid_counter_o
		if(dest in loc_id_list3):
			pass
		else:
			valid_counter_d = "Incorrect Destination Entry at : " +str(row_num)+ "."
			counter += 1
			error+=valid_counter_d
		if(time1<=24) and (time2==":") and (time3<=60):
			pass
			#print(time1,time2,time3)
			#valid_counter_t = "Time Input is Correct"
		else:
			#print("incorrect input time format, please enter as a string 17:30")
			valid_counter_t = "Incorrect time format at row : "+str(row_num)+"."
			counter += 1
			error+= valid_counter_t
		if(week in list_days_week):
			pass
			#valid_counter_w = "Weekday Input is Correct"
		else:
			#print("FALSE")
			valid_counter_w = "Weekday Input is incorrect at row : " +str(row_num)+ "."
			counter += 1
			error+=valid_counter_w
		

	if(counter==0):
		print("Form Accepted")
		#return("Form Accepted")
	else:
		print('Number of Invalid Cell Entries is: '+str(counter)+'.'+str(error)+"Please enter time in 'HH:MM' format. Eg.17:30.Please mention weekday as 'Ddd' format. Eg'Mon', 'Tue', etc.")
		#print(week)
		#try:
	
	
	if(counter==0):
		fp_out = openpyxl.Workbook()
		sheet_out = fp_out.active
		k=0
		for j in range(num_loc):
			row_num = loc_id_list[j][0]
			origin = loc_id_list[j][1]
			destination = loc_id_list[j][2]
			time = loc_id_list[j][3]
			week = loc_id_list[j][4]
			#print(origin,destination)
			
			if(origin in list_NPH) and (destination in list_NPH):
					sheet_out.cell(row = k+2 , column = 1).value = row_num
					sheet_out.cell(row = k+2 , column = 2).value = origin
					sheet_out.cell(row = k+2 , column = 3).value = destination
					sheet_out.cell(row = k+2 , column = 4).value = time
					sheet_out.cell(row = k+2 , column = 5).value = week
					k+=1
			else:
				pass
		
					#print("TRUE")
					#valid_counter_o = "Origin Input is Correct"
		fp_out.save('C:\\web2py_win\\web2py\\applications\\IndiaPost_UI\\Source Code\\IndiaPost\\NPH to NPH.xlsx')
		fp_out.close()

		fp_out = openpyxl.Workbook()
		sheet_out = fp_out.active
		k=0
		for j in range(num_loc):
			row_num = loc_id_list[j][0]
			origin = loc_id_list[j][1]
			destination = loc_id_list[j][2]
			time = loc_id_list[j][3]
			week = loc_id_list[j][4]
			#print(origin,destination)
			
			if(origin in list_NPH) and (destination in list_SPH):
					sheet_out.cell(row = k+2 , column = 1).value = row_num
					sheet_out.cell(row = k+2 , column = 2).value = origin
					sheet_out.cell(row = k+2 , column = 3).value = destination
					sheet_out.cell(row = k+2 , column = 4).value = time
					sheet_out.cell(row = k+2 , column = 5).value = week
					k+=1
			else:
				pass
					#print("TRUE")
					#valid_counter_o = "Origin Input is Correct"
		fp_out.save('C:\\web2py_win\\web2py\\applications\\IndiaPost_UI\\Source Code\\IndiaPost\\NPH to SPH.xlsx')
		fp_out.close()

		fp_out = openpyxl.Workbook()
		sheet_out = fp_out.active
		k=0
		for j in range(num_loc):
			row_num = loc_id_list[j][0]
			origin = loc_id_list[j][1]
			destination = loc_id_list[j][2]
			time = loc_id_list[j][3]
			week = loc_id_list[j][4]
			#print(origin,destination)
			
			if(origin in list_SPH) and (destination in list_NPH):
					sheet_out.cell(row = k+2 , column = 1).value = row_num
					sheet_out.cell(row = k+2 , column = 2).value = origin
					sheet_out.cell(row = k+2 , column = 3).value = destination
					sheet_out.cell(row = k+2 , column = 4).value = time
					sheet_out.cell(row = k+2 , column = 5).value = week
					k+=1
			else:
				pass
					#print("TRUE")
					#valid_counter_o = "Origin Input is Correct"
		fp_out.save('C:\\web2py_win\\web2py\\applications\\IndiaPost_UI\\Source Code\\IndiaPost\\SPH to NPH.xlsx')
		fp_out.close()

		fp_out = openpyxl.Workbook()
		sheet_out = fp_out.active
		k=0
		for j in range(num_loc):
			row_num = loc_id_list[j][0]
			origin = loc_id_list[j][1]
			destination = loc_id_list[j][2]
			time = loc_id_list[j][3]
			week = loc_id_list[j][4]
			#print(origin,destination)
			
			if(origin in list_SPH) and (destination in list_SPH):
					sheet_out.cell(row = k+2 , column = 1).value = row_num
					sheet_out.cell(row = k+2 , column = 2).value = origin
					sheet_out.cell(row = k+2 , column = 3).value = destination
					sheet_out.cell(row = k+2 , column = 4).value = time
					sheet_out.cell(row = k+2 , column = 5).value = week
					k+=1
			else:
				pass
					#print("TRUE")
					#valid_counter_o = "Origin Input is Correct"
		fp_out.save('C:\\web2py_win\\web2py\\applications\\IndiaPost_UI\\Source Code\\IndiaPost\\SPH to SPH.xlsx')
		fp_out.close()




	else:
		pass
if __name__ == '__main__':
	main()
'''
		#elif(origin in list_NPH):
			#print("FALSE")
		#	valid_counter_o = "Origin in not an NPH."
			counter += 1
		if(destination in list_NPH):
			pass
			#print("TRUE")
			#valid_counter_d = "Destination Input is Correct"
		else:
			#print("FALSE")
			valid_counter_d = "Destination in not an NPH."
			counter += 1
		
		#return(valid_counter_o,valid_counter_d,valid_counter_t,valid_counter_w)
'''
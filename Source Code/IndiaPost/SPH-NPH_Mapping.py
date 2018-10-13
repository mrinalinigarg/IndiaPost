import csv
from datetime import datetime, timedelta
import openpyxl
import pandas as pd
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.styles import PatternFill
import xlsxwriter
from openpyxl import Workbook


curDir = "C:\\web2py_win\\web2py\\applications\\IndiaPost_UI\\input files\\"
pwd = "C:\\web2py_win\\web2py\\applications\\IndiaPost_UI\\Source Code\\IndiaPost\\"
df_loc = pd.read_excel(curDir+'Location_Details.xlsx') #Reading the Values of 369 locations
loc_id_list = df_loc.values.tolist() #Creating a list loc_id_list for 369 locations
num_loc = len(loc_id_list)
print(num_loc)

#print(loc_id_list[0][1],loc_id_list[0][5], loc_id_list[0][6] )

#sheet_out.title =  "Sheet1"
'''
fp_out = openpyxl.load_workbook(pwd+ "Test.xlsx")

sheet_out = fp_out.active
sheet_out.cell(row = 1, column = 1).value = "S.No."
sheet_out.cell(row = 1, column = 2).value = "SPH Location"
sheet_out.cell(row = 1, column = 3).value = "NPH Mapping"
'''
data_list = [[] for x in range(num_loc)]
data_list[0].append("S.No.")
data_list[0].append("SPH Location")
data_list[0].append("NPH Mapping")
sno=1



for i in range(num_loc):
	if(loc_id_list[i][5]=="SPH"):
		'''sheet_out.cell(row = sno+1, column = 1).value = sno
		sheet_out.cell(row = sno+1, column = 2).value=loc_id_list[i][1]
		sheet_out.cell(row = sno+1, column = 3).value=loc_id_list[i][6]
		print(sno,loc_id_list[i][1],loc_id_list[i][6])
		'''
		data_list[sno].append(sno)
		data_list[sno].append(loc_id_list[i][1])
		data_list[sno].append(loc_id_list[i][6])
		sno+=1
	else:
		pass


	

with open(pwd+'SPH-NPH Mapping.csv', 'w', newline = '') as fp_out: #Writing the values as list of lists in 1 go to decrease time complexity
	writer = csv.writer(fp_out)
	writer.writerows(data_list)

wb = Workbook()
ws = wb.active
with open(pwd+'SPH-NPH Mapping.csv', 'r') as f:
    for row in csv.reader(f):
        ws.append(row)
wb.save(pwd+'SPH-NPH Mapping.xlsx')

#fp_out.save(pwd+"SPH-NPH Mapping.xlsx")
import pandas as pd
import csv
import openpyxl

def main():

	curDir = "C:\\web2py_win\\web2py\\applications\\IndiaPost_UI\\input files\\"
	pwd = "C:\\web2py_win\\web2py\\applications\\IndiaPost_UI\\Source Code\\IndiaPost\\"
	df_loc = pd.read_excel(pwd+'Circle_List.xlsx') #Reading the Values of 369 locations
	loc_id_list = df_loc.values.tolist() #Creating a list loc_id_list for 369 locations
	num_loc = len(loc_id_list) #Storing the total number of locations

	ctr = 0
	data_list = [[] for x in range(num_loc*num_loc)]
	for i in range(num_loc):
		for j in range(num_loc):
			data_list[ctr].extend([loc_id_list[i][1], loc_id_list[j][1], 0])#Creating a list of list to generate 369x369 Matrix of locations and initialising with 3rd column(represents number of total flights from location A to B) as '0'
			ctr += 1

	dict_id = {}
	dict_loc_cir = {} 
	 #Creating a Dictionary to allocate Ids to 369 Locations
	for i in range(num_loc):
		dict_id[loc_id_list[i][1]] = loc_id_list[i][0] - 1
		dict_loc_cir[loc_id_list[i][1]] = loc_id_list[i][3]


	#print(dict_id)
	#print(dict_loc_cir)
	
	df_circle = pd.read_excel(pwd+'Circle_ID.xlsx')
	cir_id_list = df_circle.values.tolist()
	num_cir = len(cir_id_list)

	
	dict_cir_id = {} #Creating a Dictionary to allocate Ids to 24 Circles
	for i in range(num_cir):
		dict_cir_id[cir_id_list[i][1]] = cir_id_list[i][0]

	#print(dict_cir_id)

	df_cir_loc = pd.read_excel(pwd+'Circle_Row.xlsx')
	cir_loc_id_list = df_cir_loc.values.tolist()
	num_cir_loc = len(cir_loc_id_list)	
	
	
	dict_cir_loc_id = {} #Creating a Dictionary to allocate Ids to 24 Circles
	for i in range(num_cir_loc):
		dict_cir_loc_id[cir_loc_id_list[i][0]] = cir_loc_id_list[i][1:-1]
	
	#print(dict_cir_loc_id)
	#print(dict_id['MUMBAI'])
	
	df_flt = pd.read_excel(pwd+'Location_names.xlsx')#Reading corrected names of airport locations
	list_flt = df_flt.values.tolist()
	dict_flt = {}#Creating a Dictionary to assign corrected names to the airport locations
	for i in range(len(list_flt)):
		dict_flt[list_flt[i][0]] = list_flt[i][1]

	#print(dict_flt)
	
	fp_in = openpyxl.load_workbook(curDir +'Scheduler.xlsx')#Reading Flight Schedule for all flights
	sheet_in = fp_in.get_sheet_by_name(fp_in.get_sheet_names()[0])
	num_rows = sheet_in.max_row - 1 #Calculate total number of Rows in our Flight Schedule Sheet

	dict_err_src = {} #Initialising source Dictionary
	dict_err_dest = {} #Initialising destination Dictionary
	for i in range(num_rows):
		name_src = str(sheet_in.cell(row = i+2, column = 2).value)
		name_dest = str(sheet_in.cell(row = i+2, column = 3).value)
		print(name_src,name_dest)
		if(name_src != name_dest):
			if ((dict_flt[name_src] in dict_id) and (dict_flt[name_dest] in dict_id)): #Checking for Direct Flights from A to B
				src_id = dict_id[dict_flt[name_src]] # Unique ID of Source Location
				dest_id = dict_id[dict_flt[name_dest]]# Unique ID of Destination Location
				map_id = num_loc*src_id + dest_id
				data_list[map_id][2] += 1 #Counter for Total Number of Flights from A to B 
				mon = int(sheet_in.cell(row = i+2, column = 4).value)
				tue = int(sheet_in.cell(row = i+2, column = 5).value)
				wed = int(sheet_in.cell(row = i+2, column = 6).value)
				thu = int(sheet_in.cell(row = i+2, column = 7).value)
				fri = int(sheet_in.cell(row = i+2, column = 8).value)
				sat = int(sheet_in.cell(row = i+2, column = 9).value)
				sun = int(sheet_in.cell(row = i+2, column = 10).value)

				week_scd = (1*mon)+(2*tue)+(4*wed)+(8*thu)+(16*fri)+(32*sat)+(64*sun)


				time_dep = str(sheet_in.cell(row = i+2, column = 11).value)
				time_arr = str(sheet_in.cell(row = i+2, column = 12).value)
				#day_arr = str(sheet_in.cell(row = i+2, column = 13).value)
				transit_time = str(sheet_in.cell(row = i+2, column = 13).value)
				
				#Converting time to 4 Digit Strings to ensure a consistant nomenclature
				if(len(time_dep) == 1):
					time_dep = '000' + time_dep			
				elif(len(time_dep) == 2):
					time_dep = '00' + time_dep
				elif(len(time_dep) == 3):
					time_dep = '0' + time_dep
				
				if(len(time_arr) == 1):
					time_arr = '000' + time_arr			
				elif(len(time_arr) == 2):
					time_arr = '00' + time_arr
				elif(len(time_arr) == 3):
					time_arr = '0' + time_arr
				
				#data_list[map_id].extend([week_scd, time_dep, time_arr,transit_time,mode])#Appending the Values for all flights from A to B in the row = map-id
				data_list[map_id].extend([week_scd, time_dep, time_arr,transit_time])
			else:
				print(dict_flt[name_src], dict_flt[name_dest]) #Print this statement if Source or Destination is entered incorrectly or does not exist in dict_id dictionary

	with open(pwd+'output_train.csv', 'w', newline = '') as fp_out: #Writing the values as list of lists in 1 go to decrease time complexity
		writer = csv.writer(fp_out)
		writer.writerows(data_list)
	
	return 
	#End of main()


if __name__ == "__main__":
	main()

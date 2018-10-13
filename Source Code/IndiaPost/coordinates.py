import os
import pandas as pd
from openpyxl import load_workbook
import operator
import csv
from time import *
from datetime import *
import openpyxl
import operator
from vincenty import vincenty
#from geopy.distance import vincenty

curDir = "C:\\web2py_win\\web2py\\applications\\IndiaPost_UI\\input files\\"
df_loc = pd.read_excel(curDir+'Location_Details.xlsx') #Reading the Values of 369 locations
loc_id_list = df_loc.values.tolist() #Creating a list loc_id_list for 369 locations
num_loc = len(loc_id_list)

buff_in = openpyxl.load_workbook(curDir+'Buffer.xlsx') # buffer input automation
sheet_in_buff = buff_in.get_sheet_by_name('Sheet1')

surf_speed = int(sheet_in_buff.cell(row = 2, column = 9).value)

data_list = [[] for x in range(num_loc*num_loc)]
for i in range(num_loc):
	for j in range(num_loc):
		lat_i = float(loc_id_list[i][2])
		long_i = float(loc_id_list[i][3])
		lat_j = float(loc_id_list[j][2])
		long_j = float(loc_id_list[j][3])

		city_1_name = str(loc_id_list[i][1])
		city_2_name = str(loc_id_list[j][1])

		city_1 = (lat_i,long_i)
		city_2 = (lat_j,long_j)

		distance = vincenty(city_1,city_2)
		Time = distance/surf_speed

		if (distance<=50) and (distance!=0):
			Time += 0.5
			distance+=20
		elif(distance>50) and (distance<=100):
			Time += 1.5
			distance+=40
		elif(distance>100) and (distance<=200):
			Time += 2.5
			distance+=90
		elif(distance>200) and (distance<=500):
			Time += 5
			distance+=180
		elif(distance>500) and (distance<=1000):
			Time += 7
			distance+=350
		elif(distance>1000) and (distance<=1500):
			Time += 10
			distance+=500
		elif(distance>1500) and (distance<=2000):
			Time += 14
			distance+=800
		elif(distance>2000) and (distance<=3000):
			Time += 18
			distance+=1200
		elif(distance>3000):
			Time += 24
			distance+=1700



		#print(city_1_name,city_2_name,distance)
		
		map_id = i*num_loc+j
		#print(map_id)

		data_list[map_id].append(city_1_name)
		data_list[map_id].append(city_2_name)
		data_list[map_id].append(distance)
		data_list[map_id].append(Time)

with open('C:\\web2py_win\\web2py\\applications\\IndiaPost_UI\\Source Code\\IndiaPost\\coordinates.csv', 'w', newline = '') as fp_out: #Writing the values as list of lists in 1 go to decrease time complexity
	writer = csv.writer(fp_out)
	writer.writerows(data_list)

with open('C:\\web2py_win\\web2py\\applications\\IndiaPost_UI\\Source Code\\IndiaPost\\coordinates.csv', 'r') as fp_in:
	reader = csv.reader(fp_in)
	list_loc_to_id = list(reader)




data_list_matrix = [[0 for col in range(num_loc)] for row in range(num_loc)]
#print(data_list_matrix)
for k in range(num_loc):
	for l in range(num_loc):
		row = (k*num_loc)+l
		#print(list_loc_to_id[row][2])
		data_list_matrix[k][l] = list_loc_to_id[row][2]
		
print(data_list_matrix)

#with open('C:\\web2py_win\\web2py\\applications\\IndiaPost_UI\\Source Code\\IndiaPost\\coordinates.csv', 'w', newline = '') as fp_out: #Writing the values as list of lists in 1 go to decrease time complexity
with open('C:\\web2py_win\\web2py\\applications\\IndiaPost_UI\\Source Code\\IndiaPost\\surf_matrix.csv', 'w', newline = '') as fp_out: #Writing the values as list of lists in 1 go to decrease time complexity
	writer = csv.writer(fp_out)
	writer.writerows(data_list_matrix)

import csv
from datetime import datetime, timedelta
import openpyxl
import time
import math
import os
import pandas as pd
from openpyxl.styles import Color, Fill, Font
from openpyxl.cell import Cell
from openpyxl.styles import PatternFill
import xlsxwriter
from openpyxl.workbook import Workbook
import numpy as np
#from openpyxl.cell import get_column_letter

def main():	
	curDir = "C:\\web2py_win\\web2py\\applications\\IndiaPost_UI\\Source Code\\IndiaPost\\"
	weekday_dict={'Mon':1, 'Tue':2, 'Wed':3, 'Thu':4, 'Fri':5, 'Sat':6, 'Sun':7}

	fp_in = openpyxl.load_workbook(curDir +'consolidate.xlsx')
	sheet_in = fp_in.get_sheet_by_name('Sheet')
	num_rows = sheet_in.max_row - 1
	print(num_rows, "abc")
	
	count = np.zeros(sheet_in.max_row-3,dtype=np.int)
	#counter1=1
	#counter2=0
	for row1 in range(sheet_in.max_row-3):
		#counter1 = counter1+1
		#print(val)
		for col1 in range(sheet_in.max_column-2):
			if(str(sheet_in.cell(row = row1+2, column = col1+1).value) == "None"):
				break
			else:
				count[row1] += 1
	#print(count)
	#print(row1)	
	

	

	    
	fp_out = openpyxl.Workbook()
	sheet_out = fp_out.active

	sheet_out.cell(row = 1, column = 1).value = "Legend for Colour coding :"
	sheet_out.cell(row = 1, column = 1).font = Font(bold=True)
	sheet_out.merge_cells('A1:D1')
	sheet_out.cell(row = 2, column = 5).fill = PatternFill(fgColor='FFEE08', fill_type = 'solid')
	sheet_out.cell(row = 2, column = 6).value = "Corresponding NPH location mapped for a SPH location."
	sheet_out.cell(row = 2, column = 6).font = Font(bold=True)
	sheet_out.merge_cells('F2:J2')
	sheet_out.cell(row = 3, column = 5).fill = PatternFill(fgColor='008000', fill_type = 'solid')
	sheet_out.cell(row = 3, column = 6).value = "Total Transit Time in Hours for an O-D pair."
	sheet_out.cell(row = 3, column = 6).font = Font(bold=True)
	sheet_out.merge_cells('F3:J3')
	sheet_out.cell(row = 4, column = 5).fill = PatternFill(fgColor='000000FF', fill_type = 'solid')
	sheet_out.cell(row = 4, column = 6).value = "Transit Mode is via Air."
	sheet_out.cell(row = 4, column = 6).font = Font(bold=True)
	sheet_out.merge_cells('F4:J4')

	sheet_out.cell(row = 6, column = 1).value = "Input Details"
	sheet_out.cell(row = 6, column = 1).font = Font(bold=True)
	sheet_out.merge_cells('A6:K6')
	sheet_out.cell(row = 6, column = 12).value = "Transit Leg 1"
	sheet_out.cell(row = 6, column = 12).font = Font(bold=True)
	sheet_out.merge_cells('L6:S6')
	sheet_out.cell(row = 6, column = 20).value = "Transit Leg 2"
	sheet_out.cell(row = 6, column = 20).font = Font(bold=True)
	sheet_out.merge_cells('T6:AA6')
	sheet_out.cell(row = 6, column = 28).value = "Transit Leg 3"
	sheet_out.cell(row = 6, column = 28).font = Font(bold=True)
	sheet_out.merge_cells('AB6:AI6')
	sheet_out.cell(row = 6, column = 36).value = "Transit Leg 4"
	sheet_out.cell(row = 6, column = 36).font = Font(bold=True)
	sheet_out.merge_cells('AJ6:AQ6')
	sheet_out.cell(row = 6, column = 44).value = "Transit Leg 5"
	sheet_out.cell(row = 6, column = 44).font = Font(bold=True)
	sheet_out.merge_cells('AR6:AY6')
	sheet_out.cell(row = 6, column = 52).value = "Transit Leg 6"
	sheet_out.cell(row = 6, column = 52).font = Font(bold=True)
	sheet_out.merge_cells('AZ6:BG6')


	sheet_out.append(['Serial Number','Origin','NPH/SPH','Origin NPH','Destination','NPH/SPH','Destination NPH','Dispatch Time','Dispatch Day','Total Transit Time(in Hours)','Package Arrival Day','Departure from Origin','Day of Departure','Transit Point 1','Arrival at Transit Point 1','Day of Arrival at Transit Point 1','Arrival at Transit Point 1 with buffer','Day of Arrival at Transit Point 1 with buffer','Mode of Travel','Departure from Transit Point 1','Day of Departure from Transit Point 1','Transit Point 2','Arrival at Transit Point 2','Day of Arrival at Transit Point 2','Arrival at Transit Point 2 with buffer','Day of Arrival at Transit Point 2 with buffer','Mode of Travel','Departure from Transit Point 2','Day of Departure from Transit Point 2','Transit Point 3','Arrival at Transit Point 3','Day of Arrival at Transit Point 3','Arrival at Transit Point 3 with buffer','Day of Arrival at Transit Point 3 with buffer','Mode of Travel','Departure from Transit Point 3','Day of Departure from Transit Point 3','Transit Point 4','Arrival at Transit Point 4','Day of Arrival at Transit Point 4','Arrival at Transit Point 4 with buffer','Day of Arrival at Transit Point 4 with buffer','Mode of Travel','Departure from Transit Point 4','Day of Departure from Transit Point 4','Transit Point 5','Arrival at Transit Point 5','Day of Arrival at Transit Point 5','Arrival at Transit Point 5 with buffer','Day of Arrival at Transit Point 5 with buffer','Mode of Travel','Departure from Transit Point 5','Day of Departure from Transit Point 5','Transit Point 6','Arrival at Transit Point 6','Day of Arrival at Transit Point 6','Arrival at Transit Point 6 with buffer','Day of Arrival at Transit Point 6 with buffer','Mode of Travel'])


	for i in range(num_rows-2):

		Arr_time = (str(sheet_in.cell(row = i+2, column = (count[i] -2)).value))
		Arr_Day = (str(sheet_in.cell(row = i+2, column = (count[i] -1)).value))
		Dept_time = (str(sheet_in.cell(row = i+2, column = 4).value))
		Dept_Day = (str(sheet_in.cell(row = i+2, column = 5).value))
		#print(Arr_time,Dept_time, Arr_time.split(':'), Dept_time.split(':'), format(Arr_time))

		'''
		str_format = '%Y-%m-%d %H:%M:%S'
		datetime.strptime('Arr_time', str_format)
		datetime.strptime('Dept_time', str_format)
		'''

		
		'''
		try:
			time1a,time2a = Arr_time.split(':')[0:2]
		except ValueError:
			time1a,time2a = Arr_time.split(':')[1:3]
		try:
			time1d,time2d  = Dept_time.split(':')[0:2]
		except ValueError:
			time1d,time2d  = Dept_time.split(':')[1:3]


		'''
		if(Arr_time == '1899-12-30 00:00:00'):
			time1a = 00
			time2a = 00
		else:
			time1a,time2a = Arr_time.split(':')[0:2]
		
		if(Dept_time == '1899-12-30 00:00:00'):
			time1d = 00
			time2d = 00
		else:
			time1d,time2d  = Dept_time.split(':')[0:2]

		

		time1a = int(time1a)
		time2a = int(time2a)
		time1d = int(time1d)
		time2d = int(time2d)

		print(time1a,time2a)
		print(time1d,time2d)
		'''

		time1a = Arr_time.hour
		time2a = Arr_time.minute
		time1d = Dept_time.hour
		time2d = Dept_time.minute

		print(time1a,time2a)
		print(time1d,time2d)
		time1a = int(Arr_time[0:2])
		time2a = str(Arr_time[2])
		time3a = int(Arr_time[3:5])

		time1d = int(Dept_time[0:2])
		time2d = str(Dept_time[2])
		time3d = int(Dept_time[3:5])


		
		

		'''




		arrival_datetime = datetime(1,1,weekday_dict[Arr_Day],time1a,time2a)
		dispath_datetime = datetime(1,1,weekday_dict[Dept_Day],time1d,time2d)

		#time_diff = timedelta.total_seconds(arrival_datetime - dispath_datetime)
		#time_diff_day = math.ceil(time_diff/(3600*24))

		time_diff = (weekday_dict[Arr_Day] - weekday_dict[Dept_Day])
		time_diff_day = math.ceil(time_diff) + 1

		if (time1a>7) and (time2a>0):
			print(time1a, time2a)
			time_diff_day1 = time_diff_day+1
		else:
			time_diff_day1 = time_diff_day


		print(time_diff_day,time_diff_day1)
			
		sr=str(sheet_in.cell(row = i+2, column = 1).value)
		print(sr)
		sheet_out.cell(row = i+8, column = 1).value = int(sheet_in.cell(row = i+2, column = 1).value)
		sheet_out.cell(row = i+8, column = 2).value = str(sheet_in.cell(row = i+2, column = 2).value)
		sheet_out.cell(row = i+8, column = 3).value = str(sheet_in.cell(row = i+2, column = 57).value)
		b=str(sheet_in.cell(row = i+2, column = 6).value)
		#print(sheet_out.cell(row = i+8, column = 5).value)
		sheet_out.cell(row = i+8, column = 4).value = str(sheet_in.cell(row = i+2, column = 6).value)
		if(b!='-'):
			sheet_out.cell(row = i+8, column = 4).fill = PatternFill(fgColor='FFEE08', fill_type = 'solid')
		c=str(sheet_in.cell(row = i+2, column = 7).value)
		sheet_out.cell(row = i+8, column = 7).value = str(sheet_in.cell(row = i+2, column = 7).value)
		if(c!='-'):
			sheet_out.cell(row = i+8, column = 7).fill = PatternFill(fgColor='FFEE08', fill_type = 'solid')
		sheet_out.cell(row = i+8, column = 5).value = str(sheet_in.cell(row = i+2, column = 3).value)
		sheet_out.cell(row = i+8, column = 6).value = str(sheet_in.cell(row = i+2, column = 58).value)
		a=float(sheet_in.cell(row = i+2, column = 8).value)
		sheet_out.cell(row = i+8, column = 10).value = format(a, '.2f')
		sheet_out.cell(row = i+8, column = 10).fill = PatternFill(fgColor='008000', fill_type = 'solid')
		sheet_out.cell(row = i+8, column = 10).font = Font(bold=True)
		sheet_out.cell(row = i+8, column = 8).value = str(sheet_in.cell(row = i+2, column = 4).value)
		sheet_out.cell(row = i+8, column = 9).value = str(sheet_in.cell(row = i+2, column = 5).value)
		sheet_out.cell(row = i+8, column = 11).value=time_diff_day1
		sheet_out.cell(row = i+8, column = 12).value = str(sheet_in.cell(row = i+2, column = 9).value)
		sheet_out.cell(row = i+8, column = 13).value = str(sheet_in.cell(row = i+2, column = 10).value)
		sheet_out.cell(row = i+8, column = 14).value = str(sheet_in.cell(row = i+2, column = 11).value)
		sheet_out.cell(row = i+8, column = 15).value = str(sheet_in.cell(row = i+2, column = 12).value)#
		sheet_out.cell(row = i+8, column = 16).value = str(sheet_in.cell(row = i+2, column = 13).value)
		sheet_out.cell(row = i+8, column = 17).value = str(sheet_in.cell(row = i+2, column = 14).value)
		sheet_out.cell(row = i+8, column = 18).value = str(sheet_in.cell(row = i+2, column = 15).value)
		sheet_out.cell(row = i+8, column = 19).value = str(sheet_in.cell(row = i+2, column = 16).value)
		sheet_out.cell(row = i+8, column = 20).value = str(sheet_in.cell(row = i+2, column = 17).value)
		sheet_out.cell(row = i+8, column = 21).value = str(sheet_in.cell(row = i+2, column = 18).value)
		sheet_out.cell(row = i+8, column = 22).value = str(sheet_in.cell(row = i+2, column = 19).value)
		sheet_out.cell(row = i+8, column = 23).value = str(sheet_in.cell(row = i+2, column = 20).value)
		sheet_out.cell(row = i+8, column = 24).value = str(sheet_in.cell(row = i+2, column = 21).value)
		sheet_out.cell(row = i+8, column = 25).value = str(sheet_in.cell(row = i+2, column = 22).value)
		sheet_out.cell(row = i+8, column = 26).value = str(sheet_in.cell(row = i+2, column = 23).value)
		sheet_out.cell(row = i+8, column = 27).value = str(sheet_in.cell(row = i+2, column = 24).value)
		sheet_out.cell(row = i+8, column = 28).value = str(sheet_in.cell(row = i+2, column = 25).value)
		sheet_out.cell(row = i+8, column = 29).value = str(sheet_in.cell(row = i+2, column = 26).value)
		sheet_out.cell(row = i+8, column = 30).value = str(sheet_in.cell(row = i+2, column = 27).value)
		sheet_out.cell(row = i+8, column = 31).value = str(sheet_in.cell(row = i+2, column = 28).value)
		sheet_out.cell(row = i+8, column = 32).value = str(sheet_in.cell(row = i+2, column = 29).value)
		sheet_out.cell(row = i+8, column = 33).value = str(sheet_in.cell(row = i+2, column = 30).value)
		sheet_out.cell(row = i+8, column = 34).value = str(sheet_in.cell(row = i+2, column = 31).value)
		sheet_out.cell(row = i+8, column = 35).value = str(sheet_in.cell(row = i+2, column = 32).value)
		sheet_out.cell(row = i+8, column = 36).value = str(sheet_in.cell(row = i+2, column = 33).value)
		sheet_out.cell(row = i+8, column = 37).value = str(sheet_in.cell(row = i+2, column = 34).value)
		sheet_out.cell(row = i+8, column = 38).value = str(sheet_in.cell(row = i+2, column = 35).value)
		sheet_out.cell(row = i+8, column = 39).value = str(sheet_in.cell(row = i+2, column = 36).value)
		sheet_out.cell(row = i+8, column = 40).value = str(sheet_in.cell(row = i+2, column = 37).value)
		sheet_out.cell(row = i+8, column = 41).value = str(sheet_in.cell(row = i+2, column = 38).value)
		sheet_out.cell(row = i+8, column = 42).value = str(sheet_in.cell(row = i+2, column = 39).value)
		sheet_out.cell(row = i+8, column = 43).value = str(sheet_in.cell(row = i+2, column = 40).value)
		sheet_out.cell(row = i+8, column = 44).value = str(sheet_in.cell(row = i+2, column = 41).value)
		sheet_out.cell(row = i+8, column = 45).value = str(sheet_in.cell(row = i+2, column = 42).value)
		sheet_out.cell(row = i+8, column = 46).value = str(sheet_in.cell(row = i+2, column = 43).value)
		sheet_out.cell(row = i+8, column = 47).value = str(sheet_in.cell(row = i+2, column = 44).value)
		sheet_out.cell(row = i+8, column = 48).value = str(sheet_in.cell(row = i+2, column = 45).value)
		sheet_out.cell(row = i+8, column = 49).value = str(sheet_in.cell(row = i+2, column = 46).value)
		sheet_out.cell(row = i+8, column = 50).value = str(sheet_in.cell(row = i+2, column = 47).value)
		sheet_out.cell(row = i+8, column = 51).value = str(sheet_in.cell(row = i+2, column = 48).value)
		sheet_out.cell(row = i+8, column = 52).value = str(sheet_in.cell(row = i+2, column = 49).value)
		sheet_out.cell(row = i+8, column = 53).value = str(sheet_in.cell(row = i+2, column = 50).value)
		sheet_out.cell(row = i+8, column = 54).value = str(sheet_in.cell(row = i+2, column = 51).value)
		sheet_out.cell(row = i+8, column = 55).value = str(sheet_in.cell(row = i+2, column = 52).value)
		sheet_out.cell(row = i+8, column = 56).value = str(sheet_in.cell(row = i+2, column = 53).value)
		sheet_out.cell(row = i+8, column = 57).value = str(sheet_in.cell(row = i+2, column = 54).value)
		sheet_out.cell(row = i+8, column = 58).value = str(sheet_in.cell(row = i+2, column = 55).value)
		sheet_out.cell(row = i+8, column = 59).value = str(sheet_in.cell(row = i+2, column = 56).value)
	




	#num_rows = num_rows_nn+num_rows_ns+num_rows_sn+num_rows_ss
	for j in range(num_rows):
		for k in range(59):
			mode = str(sheet_out.cell(row = j+8, column = k+1).value)
			if(mode=='Air'):
				sheet_out.cell(row = j+8, column = k+1).fill = PatternFill(fgColor='000000FF', fill_type = 'solid')
			if(mode=='None'):
				sheet_out.cell(row = j+8, column = k+1).value = ""

	c = sheet_out['L8']
	sheet_out.freeze_panes = c

	for cell in sheet_out["7:7"]:
		cell.font = Font(bold=True)
	for cell in sheet_out["6:6"]:
		cell.font = Font(bold=True)

	fp_out.save('C:\\web2py_win\\web2py\\applications\\IndiaPost_UI\\private\\Output.xlsx')

if __name__ == '__main__':
	main()
		
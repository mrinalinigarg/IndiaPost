import pandas as pd
import csv
import openpyxl
from datetime import datetime, timedelta
'''
df_loc = pd.read_excel('Rail_input format.xlsx', 'Sheet1') #Reading the Values of 369 locations
loc_id_list = df_loc.values.tolist() #Creating a list loc_id_list for 369 locations
num_loc = len(loc_id_list)
'''
curDir = "C:\\web2py_win\\web2py\\applications\\IndiaPost_UI\\input files\\"
fp_in = openpyxl.load_workbook(curDir+'Rail_schedule_format.xlsx', 'Sheet1')#Reading Flight Schedule for all flights
sheet_in = fp_in.get_sheet_by_name(fp_in.get_sheet_names()[0])
num_rows = sheet_in.max_row - 1
num_cols = sheet_in.max_column - 1

print(num_rows,num_cols)
'''
for i in range(num_rows):
	for j in range(num_cols):
		value= str(sheet_in.cell(row = 1, column = j+1).value)
		print(value)
'''

fp_out = openpyxl.Workbook()
sheet_out = fp_out.active

sheet_out.append(['Serial No.','Origin','Destination','Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday','Departure Time','Arrival Time','Day of Arrival','Duration of travel (h)', 'Train No.'])

Sno=1
for i in range(3,num_rows):
	for j in range(4,num_cols-2,6):
		origin= str(sheet_in.cell(row = i+1, column = j).value)
		#print(origin)
		mode = str(sheet_in.cell(row = i+1, column = j+1).value)
		#print(mode)
		departure_time = str(sheet_in.cell(row = i+1, column = j+2).value)
		#print(departure_time)
		destination = str(sheet_in.cell(row = i+1, column = j+3).value)
		#print(destination)
		arrival_time = str(sheet_in.cell(row = i+1, column = j+4).value)
		#print(arrival_time)
		day_of_arrival =str(sheet_in.cell(row = i+1, column = j+5).value)
		#print(day_of_arrival)

		transit_time = 0

		#if(origin is None) or (destination is None) or (arrival_time is None) or (departure_time is None):
		#				pass
		if (origin != 'NA') and (destination != 'NA') and (arrival_time != 'NA') and (departure_time != 'NA') and (arrival_time != 'None') and (departure_time != 'None') and (day_of_arrival != 'NA') and (day_of_arrival != 'None'):

			transit_time = 0
			print(Sno,departure_time,arrival_time)

			timedep1 = int(departure_time[0:2])
			timedep2 = int(departure_time[2:4])
			time_arr1 = int(arrival_time[0:2])
			time_arr2 = int(arrival_time[2:4])

			dayar= int(day_of_arrival)

			

			if(time_arr1-timedep1>0):
				transit_time = datetime(1, 1, dayar+1, time_arr1, time_arr2) - datetime(1, 1, 1, timedep1, timedep2)
			else:
				transit_time = datetime(1, 1, dayar+1, time_arr1, time_arr2) - datetime(1, 1, 1, timedep1, timedep2)
			
			sheet_out.cell(row = Sno+1, column = 1).value = Sno
			sheet_out.cell(row = Sno+1, column = 2).value = origin
			sheet_out.cell(row = Sno+1, column = 3).value = destination
			sheet_out.cell(row = Sno+1, column = 4).value = 1
			sheet_out.cell(row = Sno+1, column = 5).value = 1
			sheet_out.cell(row = Sno+1, column = 6).value = 1
			sheet_out.cell(row = Sno+1, column = 7).value = 1
			sheet_out.cell(row = Sno+1, column = 8).value = 1
			sheet_out.cell(row = Sno+1, column = 9).value = 1
			sheet_out.cell(row = Sno+1, column = 10).value = 1
			sheet_out.cell(row = Sno+1, column = 11).value = departure_time
			sheet_out.cell(row = Sno+1, column = 12).value = arrival_time
			sheet_out.cell(row = Sno+1, column = 13).value = day_of_arrival
			sheet_out.cell(row = Sno+1, column = 14).value = transit_time.total_seconds()/3600
			sheet_out.cell(row = Sno+1, column = 15).value = mode
			Sno+=1
		else:
			print(i,j)
			pass

fp_out.save('scheduler.xlsx')
import csv
from datetime import datetime, timedelta
import openpyxl
import time
import math
import os
import pandas as pd


def main():

	curDir = "C:\\web2py_win\\web2py\\applications\\IndiaPost_UI\\input files\\"
	pwd = "C:\\web2py_win\\web2py\\applications\\IndiaPost_UI\\Source Code\\IndiaPost\\"
	#df_loc = pd.read_excel('Locations.xlsx', 'Circle_List') #Reading the Values of 369 locations
	df_loc = pd.read_excel(pwd+'Circle_List.xlsx')
	loc_id_list = df_loc.values.tolist() #Creating a list loc_id_list for 369 locations
	num_loc = len(loc_id_list) #Storing the total number of locations

	ctr = 0
	data_list = [[] for x in range(num_loc*num_loc)]
	for i in range(num_loc):
		for j in range(num_loc):
			data_list[ctr].extend([loc_id_list[i][1], loc_id_list[j][1], 0])#Creating a list of list to generate 369x369 Matrix of locations and initialising with 3rd column(represents number of total flights from location A to B) as '0'
			ctr += 1

	dict_id = {}
	dict_loc_cir = {} 
	 #Creating a Dictionary to allocate Ids to 369 Locations
	for i in range(num_loc):
		dict_id[loc_id_list[i][1]] = loc_id_list[i][0] - 1
		dict_loc_cir[loc_id_list[i][1]] = loc_id_list[i][3]


	#print(dict_id)
	#print(dict_loc_cir)

	df_circle = pd.read_excel(pwd+'Circle_ID.xlsx')
	cir_id_list = df_circle.values.tolist()
	num_cir = len(cir_id_list)

	
	dict_cir_id = {} #Creating a Dictionary to allocate Ids to 24 Circles
	for i in range(num_cir):
		dict_cir_id[cir_id_list[i][1]] = cir_id_list[i][0]

	#print(dict_cir_id)
	df_cir_loc = pd.read_excel(pwd+'Circle_Row.xlsx')
	cir_loc_id_list = df_cir_loc.values.tolist()
	num_cir_loc = len(cir_loc_id_list)	
	
	
	dict_cir_loc_id = {} #Creating a Dictionary to allocate Ids to 24 Circles
	for i in range(num_cir_loc):
		dict_cir_loc_id[cir_loc_id_list[i][0]] = cir_loc_id_list[i][1:-1]
	
	#print(dict_cir_loc_id)
	#print(dict_id['MUMBAI'])
   
	df_flt = pd.read_excel(pwd+'Location_names.xlsx')#Reading corrected names of airport locations
	list_flt = df_flt.values.tolist()
	dict_flt = {}#Creating a Dictionary to assign corrected names to the airport locations
	for i in range(len(list_flt)):
		dict_flt[list_flt[i][0]] = list_flt[i][1]

	fp_in = openpyxl.load_workbook(curDir+'Flight_schedule_format.xlsx')#Reading Flight Schedule for all flights
	sheet_in = fp_in.get_sheet_by_name(fp_in.get_sheet_names()[0])
	num_rows = sheet_in.max_row - 1 #Calculate total number of Rows in our Flight Schedule Sheet

	dict_err_src = {} #Initialising source Dictionary
	dict_err_dest = {} #Initialising destination Dictionary
	for i in range(num_rows):
		name_src = str(sheet_in.cell(row = i+2, column = 3).value)
		name_dest = str(sheet_in.cell(row = i+2, column = 5).value)

		print(name_src,name_dest)

		if ((dict_flt[name_src] in dict_id) and (dict_flt[name_dest] in dict_id) and (name_src!=name_dest)): #Checking for Direct Flights from A to B
			src_id = dict_id[dict_flt[name_src]] # Unique ID of Source Location
			dest_id = dict_id[dict_flt[name_dest]]# Unique ID of Destination Location
			map_id = num_loc*src_id + dest_id
			data_list[map_id][2] += 1 #Counter for Total Number of Flights from A to B 
			week_scd = int(sheet_in.cell(row = i+2, column = 13).value)
			time_dep = str(sheet_in.cell(row = i+2, column = 14).value)
			time_arr = str(sheet_in.cell(row = i+2, column = 15).value)
			flight_no = str(sheet_in.cell(row = i+2, column = 16).value)
			transit_time = 0

			timedep1 = int(time_dep[0:2])
			timedep2 = int(time_dep[2:4])
			time_arr1 = int(time_arr[0:2])
			time_arr2 = int(time_arr[2:4])

			print(timedep1,time_arr1)

			if(time_arr1>timedep1):
				transit_time = datetime(1, 1, 1, time_arr1, time_arr2) - datetime(1, 1, 1, timedep1, timedep2)
			else:
				transit_time = datetime(1, 1, 2, time_arr1, time_arr2) - datetime(1, 1, 1, timedep1, timedep2)
			#transit_time = str(sheet_in.cell(row = i+2, column = 11).value)
			#data_list[map_id].extend([week_scd, time_dep, time_arr,transit_time,flight_no])#Appending the Values for all flights from A to B in the row = map-id
			data_list[map_id].extend([week_scd, time_dep, time_arr,transit_time.total_seconds()/3600])
		else:
			print(dict_flt[name_src], dict_flt[name_dest]) #Print this statement if Source or Destination is entered incorrectly or does not exist in dict_id dictionary

	with open(pwd+'output_air.csv', 'w', newline = '') as fp_out: #Writing the values as list of lists in 1 go to decrease time complexity
		writer = csv.writer(fp_out)
		writer.writerows(data_list)

	return 
	#End of main()


if __name__ == "__main__":
	main()

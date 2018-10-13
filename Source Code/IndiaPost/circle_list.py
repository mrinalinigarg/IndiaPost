import csv
from datetime import datetime, timedelta
import openpyxl
import pandas as pd


curDir = "C:\\web2py_win\\web2py\\applications\\IndiaPost_UI\\input files\\"
pwd = "C:\\web2py_win\\web2py\\applications\\IndiaPost_UI\\Source Code\\IndiaPost\\"
df_loc = pd.read_excel(curDir+'Location_Details.xlsx') #Reading the Values of 369 locations
loc_id_list = df_loc.values.tolist() #Creating a list loc_id_list for 369 locations
num_loc = len(loc_id_list)
print(num_loc)

fp_out = openpyxl.load_workbook(pwd + "Circle_List.xlsx")

sheet_out = fp_out.active

sheet_out.title =  "Circle_List"
sheet_out.cell(row = 1, column = 1).value = "ID"
sheet_out.cell(row = 1, column = 2).value = "Name"
sheet_out.cell(row = 1, column = 3).value = "Circle Name"
sheet_out.cell(row = 1, column = 4).value = "Circle ID"
sheet_out.cell(row = 1, column = 5).value = "ID"


for i in range(num_loc):
	location_name = str(loc_id_list[i][1])
	location_id = int(loc_id_list[i][0])
	circle_name = str(loc_id_list[i][4])
	circle_id = int(loc_id_list[i][7])
	
	sheet_out.cell(row = i+2, column = 1).value = location_id
	sheet_out.cell(row = i+2, column = 2).value = location_name
	sheet_out.cell(row = i+2, column = 3).value = circle_name
	sheet_out.cell(row = i+2, column = 4).value = circle_id
	sheet_out.cell(row = i+2, column = 5).value = location_id
	#print(data_list[i+1][0],data_list[i+1][1])

	#Wprint(circle_name,location_id)	

#worksheet.write(data_list)
#workbook.close()

fp_out.save(pwd+"Circle_List.xlsx")

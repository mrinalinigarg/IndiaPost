import csv
from datetime import datetime, timedelta
import openpyxl
import time
import math
import os
import pandas as pd
from openpyxl.styles import Color, Fill, Font
from openpyxl.cell import Cell
from openpyxl.styles import PatternFill
import xlsxwriter

def main():	
	'''
	if not(os.path.exists("C:\\web2py_win\\web2py\\applications\\IndiaPost_UI\\Source Code\\IndiaPost\\Air_NPH to NPH.xlsx")):
		os.makedirs("C:\\web2py_win\\web2py\\applications\\IndiaPost_UI\\Source Code\\IndiaPost\\Air_NPH to NPH.xlsx")
	else:
		pass
	if not(os.path.exists("C:\\web2py_win\\web2py\\applications\\IndiaPost_UI\\Source Code\\IndiaPost\\Air_NPH to SPH.xlsx")):
		os.makedirs("C:\\web2py_win\\web2py\\applications\\IndiaPost_UI\\Source Code\\IndiaPost\\Air_NPH to SPH.xlsx")
	else:
		pass
	if not(os.path.exists("C:\\web2py_win\\web2py\\applications\\IndiaPost_UI\\Source Code\\IndiaPost\\Air_SPH to NPH.xlsx")):
		os.makedirs("C:\\web2py_win\\web2py\\applications\\IndiaPost_UI\\Source Code\\IndiaPost\\Air_SPH to NPH.xlsx")
	else:
		pass
	if not(os.path.exists("C:\\web2py_win\\web2py\\applications\\IndiaPost_UI\\Source Code\\IndiaPost\\Air_SPH to SPH.xlsx")):
		os.makedirs("C:\\web2py_win\\web2py\\applications\\IndiaPost_UI\\Source Code\\IndiaPost\\Air_SPH to SPH.xlsx")
	else:
		pass
	try:
		fp_in = openpyxl.load_workbook(curDir +'Air_NPH to NPH.xlsx')
		sheet_in = fp_in.get_sheet_by_name('Sheet')
		num_rows_nn = sheet_in.max_row - 1
	except:
		fp_out = openpyxl.Workbook()
		sheet_out = fp_out.active
		fp_out.save(curDir +'Air_NPH to NPH.xlsx')
		fp_out.close()

	try:
		fp_in2 = openpyxl.load_workbook(curDir +'Air_NPH to SPH.xlsx')
		sheet_in2 = fp_in2.get_sheet_by_name('Sheet')
		num_rows_ns = sheet_in2.max_row - 1
	except:
		fp_out = openpyxl.Workbook()
		sheet_out = fp_out.active
		fp_out.save(curDir +'Air_NPH to SPH.xlsx')
		fp_out.close()

	try:
		fp_in3 = openpyxl.load_workbook(curDir +'Air_SPH to NPH.xlsx')
		sheet_in3 = fp_in3.get_sheet_by_name('Sheet')
		num_rows_sn = sheet_in3.max_row - 1
	except:
		fp_out = openpyxl.Workbook()
		sheet_out = fp_out.active
		fp_out.save(curDir +'Air_SPH to NPH.xlsx')
		fp_out.close()

	try:
		fp_in4 = openpyxl.load_workbook(curDir +'Air_SPH to SPH.xlsx')
		sheet_in4 = fp_in4.get_sheet_by_name('Sheet')
		num_rows_ss = sheet_in4.max_row - 1
	except:
		fp_out = openpyxl.Workbook()
		sheet_out = fp_out.active
		fp_out.save(curDir +'Air_SPH to SPH.xlsx')
		fp_out.close()
	'''
	
	curDir = "C:\\web2py_win\\web2py\\applications\\IndiaPost_UI\\Source Code\\IndiaPost\\"

	

	fp_in = openpyxl.load_workbook(curDir +'Air_NPH to NPH.xlsx')
	sheet_in = fp_in.get_sheet_by_name('Sheet')
	num_rows_nn = sheet_in.max_row - 1

	fp_in2 = openpyxl.load_workbook(curDir +'Air_NPH to SPH.xlsx')
	sheet_in2 = fp_in2.get_sheet_by_name('Sheet')
	num_rows_ns = sheet_in2.max_row - 1

	fp_in3 = openpyxl.load_workbook(curDir +'Air_SPH to NPH.xlsx')
	sheet_in3 = fp_in3.get_sheet_by_name('Sheet')
	num_rows_sn = sheet_in3.max_row - 1

	fp_in4 = openpyxl.load_workbook(curDir +'Air_SPH to SPH.xlsx')
	sheet_in4 = fp_in4.get_sheet_by_name('Sheet')
	num_rows_ss = sheet_in4.max_row - 1

	fp_out = openpyxl.Workbook()
	sheet_out = fp_out.active

	sheet_out.append(['Serial Number','Origin','Destination','Dispatch Time','Dispatch Day','Origin NPH','Destination NPH','Total Transit Time','Departure from Origin','Day of Departure','Transit Point 1','Arrival at Transit Point 1','Day of Arrival at Transit Point 1','Arrival at Transit Point 1 with buffer','Day of Arrival at Transit Point 1 with buffer','Mode of Travel','Departure from Transit Point 1','Day of Departure from Transit Point 1','Transit Point 2','Arrival at Transit Point 2','Day of Arrival at Transit Point 2','Arrival at Transit Point 2 with buffer','Day of Arrival at Transit Point 2 with buffer','Mode of Travel','Departure from Transit Point 2','Day of Departure from Transit Point 2','Transit Point 3','Arrival at Transit Point 3','Day of Arrival at Transit Point 3','Arrival at Transit Point 3 with buffer','Day of Arrival at Transit Point 3 with buffer','Mode of Travel','Departure from Transit Point 3','Day of Departure from Transit Point 3','Transit Point 4','Arrival at Transit Point 4','Day of Arrival at Transit Point 4','Arrival at Transit Point 4 with buffer','Day of Arrival at Transit Point 4 with buffer','Mode of Travel','Departure from Transit Point 4','Day of Departure from Transit Point 4','Transit Point 5','Arrival at Transit Point 5','Day of Arrival at Transit Point 5','Arrival at Transit Point 5 with buffer','Day of Arrival at Transit Point 5 with buffer','Mode of Travel','Departure from Transit Point 5','Day of Departure from Transit Point 5','Transit Point 6','Arrival at Transit Point 6','Day of Arrival at Transit Point 6','Arrival at Transit Point 6 with buffer','Day of Arrival at Transit Point 6 with buffer','Mode of Travel'])

	if(num_rows_nn>0):
		for i in range(num_rows_nn):
			serial_number = int(sheet_in.cell(row = i+2, column = 1).value)
			#print(serial_number)
			
			sheet_out.cell(row = serial_number+1, column = 1).value = serial_number
			sheet_out.cell(row = serial_number+1, column = 2).value = str(sheet_in.cell(row = i+2, column = 2).value)
			sheet_out.cell(row = serial_number+1, column = 3).value = str(sheet_in.cell(row = i+2, column = 3).value)
			sheet_out.cell(row = serial_number+1, column = 4).value = str(sheet_in.cell(row = i+2, column = 4).value)
			sheet_out.cell(row = serial_number+1, column = 5).value = str(sheet_in.cell(row = i+2, column = 5).value)
			sheet_out.cell(row = serial_number+1, column = 6).value = '-'
			sheet_out.cell(row = serial_number+1, column = 7).value = '-'
			a=float(sheet_in.cell(row = i+2, column = 6).value)
			sheet_out.cell(row = serial_number+1, column = 8).value = format(a, '.2f')
			sheet_out.cell(row = serial_number+1, column = 8).fill = PatternFill(fgColor='FFEE08', fill_type = 'solid')
			sheet_out.cell(row = serial_number+1, column = 8).font = Font(bold=True)
			sheet_out.cell(row = serial_number+1, column = 9).value = str(sheet_in.cell(row = i+2, column = 7).value)
			sheet_out.cell(row = serial_number+1, column = 10).value = str(sheet_in.cell(row = i+2, column = 8).value)
			sheet_out.cell(row = serial_number+1, column = 11).value = str(sheet_in.cell(row = i+2, column = 14).value)
			sheet_out.cell(row = serial_number+1, column = 12).value = str(sheet_in.cell(row = i+2, column = 9).value)
			sheet_out.cell(row = serial_number+1, column = 13).value = str(sheet_in.cell(row = i+2, column = 10).value)
			sheet_out.cell(row = serial_number+1, column = 14).value = str(sheet_in.cell(row = i+2, column = 11).value)
			sheet_out.cell(row = serial_number+1, column = 15).value = str(sheet_in.cell(row = i+2, column = 12).value)
			sheet_out.cell(row = serial_number+1, column = 16).value = str(sheet_in.cell(row = i+2, column = 13).value)
			sheet_out.cell(row = serial_number+1, column = 17).value = str(sheet_in.cell(row = i+2, column = 15).value)
			sheet_out.cell(row = serial_number+1, column = 18).value = str(sheet_in.cell(row = i+2, column = 16).value)
			sheet_out.cell(row = serial_number+1, column = 19).value = str(sheet_in.cell(row = i+2, column = 22).value)
			sheet_out.cell(row = serial_number+1, column = 20).value = str(sheet_in.cell(row = i+2, column = 17).value)
			sheet_out.cell(row = serial_number+1, column = 21).value = str(sheet_in.cell(row = i+2, column = 18).value)
			sheet_out.cell(row = serial_number+1, column = 22).value = str(sheet_in.cell(row = i+2, column = 19).value)
			sheet_out.cell(row = serial_number+1, column = 23).value = str(sheet_in.cell(row = i+2, column = 20).value)
			sheet_out.cell(row = serial_number+1, column = 24).value = str(sheet_in.cell(row = i+2, column = 21).value)
			sheet_out.cell(row = serial_number+1, column = 25).value = str(sheet_in.cell(row = i+2, column = 23).value)
			sheet_out.cell(row = serial_number+1, column = 26).value = str(sheet_in.cell(row = i+2, column = 24).value)
			sheet_out.cell(row = serial_number+1, column = 27).value = str(sheet_in.cell(row = i+2, column = 30).value)
			sheet_out.cell(row = serial_number+1, column = 28).value = str(sheet_in.cell(row = i+2, column = 25).value)
			sheet_out.cell(row = serial_number+1, column = 29).value = str(sheet_in.cell(row = i+2, column = 26).value)
			sheet_out.cell(row = serial_number+1, column = 30).value = str(sheet_in.cell(row = i+2, column = 27).value)
			sheet_out.cell(row = serial_number+1, column = 31).value = str(sheet_in.cell(row = i+2, column = 28).value)
			sheet_out.cell(row = serial_number+1, column = 32).value = str(sheet_in.cell(row = i+2, column = 29).value)
			sheet_out.cell(row = serial_number+1, column = 33).value = str(sheet_in.cell(row = i+2, column = 31).value)
			sheet_out.cell(row = serial_number+1, column = 34).value = str(sheet_in.cell(row = i+2, column = 32).value)
			sheet_out.cell(row = serial_number+1, column = 35).value = str(sheet_in.cell(row = i+2, column = 38).value)
			sheet_out.cell(row = serial_number+1, column = 36).value = str(sheet_in.cell(row = i+2, column = 33).value)
			sheet_out.cell(row = serial_number+1, column = 37).value = str(sheet_in.cell(row = i+2, column = 34).value)
			sheet_out.cell(row = serial_number+1, column = 38).value = str(sheet_in.cell(row = i+2, column = 35).value)
			sheet_out.cell(row = serial_number+1, column = 39).value = str(sheet_in.cell(row = i+2, column = 36).value)
			sheet_out.cell(row = serial_number+1, column = 40).value = str(sheet_in.cell(row = i+2, column = 37).value)
			sheet_out.cell(row = serial_number+1, column = 41).value = str(sheet_in.cell(row = i+2, column = 39).value)
			sheet_out.cell(row = serial_number+1, column = 42).value = str(sheet_in.cell(row = i+2, column = 40).value)
			sheet_out.cell(row = serial_number+1, column = 43).value = str(sheet_in.cell(row = i+2, column = 46).value)
			sheet_out.cell(row = serial_number+1, column = 44).value = str(sheet_in.cell(row = i+2, column = 41).value)
			sheet_out.cell(row = serial_number+1, column = 45).value = str(sheet_in.cell(row = i+2, column = 42).value)
			sheet_out.cell(row = serial_number+1, column = 46).value = str(sheet_in.cell(row = i+2, column = 43).value)
			sheet_out.cell(row = serial_number+1, column = 47).value = str(sheet_in.cell(row = i+2, column = 44).value)
			sheet_out.cell(row = serial_number+1, column = 48).value = str(sheet_in.cell(row = i+2, column = 45).value)
			sheet_out.cell(row = serial_number+1, column = 49).value = str(sheet_in.cell(row = i+2, column = 47).value)
			sheet_out.cell(row = serial_number+1, column = 50).value = str(sheet_in.cell(row = i+2, column = 48).value)
			sheet_out.cell(row = serial_number+1, column = 51).value = str(sheet_in.cell(row = i+2, column = 54).value)
			sheet_out.cell(row = serial_number+1, column = 52).value = str(sheet_in.cell(row = i+2, column = 49).value)
			sheet_out.cell(row = serial_number+1, column = 53).value = str(sheet_in.cell(row = i+2, column = 50).value)
			sheet_out.cell(row = serial_number+1, column = 54).value = str(sheet_in.cell(row = i+2, column = 51).value)
			sheet_out.cell(row = serial_number+1, column = 55).value = str(sheet_in.cell(row = i+2, column = 52).value)
			sheet_out.cell(row = serial_number+1, column = 56).value = str(sheet_in.cell(row = i+2, column = 53).value)
			sheet_out.cell(row = serial_number+1, column = 57).value = "NPH"
			sheet_out.cell(row = serial_number+1, column = 58).value = "NPH"
	else:
		pass
	if(num_rows_ns>0):
		for i in range(num_rows_ns):
			serial_number = int(sheet_in2.cell(row = i+2, column = 1).value)
			
			sheet_out.cell(row = serial_number+1, column = 1).value = serial_number
			sheet_out.cell(row = serial_number+1, column = 2).value = str(sheet_in2.cell(row = i+2, column = 2).value)
			sheet_out.cell(row = serial_number+1, column = 3).value = str(sheet_in2.cell(row = i+2, column = 3).value)
			sheet_out.cell(row = serial_number+1, column = 4).value = str(sheet_in2.cell(row = i+2, column = 4).value)
			sheet_out.cell(row = serial_number+1, column = 5).value = str(sheet_in2.cell(row = i+2, column = 5).value)
			sheet_out.cell(row = serial_number+1, column = 6).value = '-'
			sheet_out.cell(row = serial_number+1, column = 7).value = str(sheet_in2.cell(row = i+2, column = 6).value)
			sheet_out.cell(row = serial_number+1, column = 7).fill = PatternFill(fgColor='FFEE08', fill_type = 'solid')
			a=float(sheet_in2.cell(row = i+2, column = 7).value)
			sheet_out.cell(row = serial_number+1, column = 8).value = format(a, '.2f')
			sheet_out.cell(row = serial_number+1, column = 8).fill = PatternFill(fgColor='FFEE08', fill_type = 'solid')
			sheet_out.cell(row = serial_number+1, column = 8).font = Font(bold=True)
			sheet_out.cell(row = serial_number+1, column = 9).value = str(sheet_in2.cell(row = i+2, column = 8).value)
			sheet_out.cell(row = serial_number+1, column = 10).value = str(sheet_in2.cell(row = i+2, column = 9).value)
			sheet_out.cell(row = serial_number+1, column = 11).value = str(sheet_in2.cell(row = i+2, column = 15).value)
			sheet_out.cell(row = serial_number+1, column = 12).value = str(sheet_in2.cell(row = i+2, column = 10).value)
			sheet_out.cell(row = serial_number+1, column = 13).value = str(sheet_in2.cell(row = i+2, column = 11).value)
			sheet_out.cell(row = serial_number+1, column = 14).value = str(sheet_in2.cell(row = i+2, column = 12).value)
			sheet_out.cell(row = serial_number+1, column = 15).value = str(sheet_in2.cell(row = i+2, column = 13).value)
			sheet_out.cell(row = serial_number+1, column = 16).value = str(sheet_in2.cell(row = i+2, column = 14).value)
			sheet_out.cell(row = serial_number+1, column = 17).value = str(sheet_in2.cell(row = i+2, column = 16).value)
			sheet_out.cell(row = serial_number+1, column = 18).value = str(sheet_in2.cell(row = i+2, column = 17).value)
			sheet_out.cell(row = serial_number+1, column = 19).value = str(sheet_in2.cell(row = i+2, column = 23).value)
			sheet_out.cell(row = serial_number+1, column = 20).value = str(sheet_in2.cell(row = i+2, column = 18).value)
			sheet_out.cell(row = serial_number+1, column = 21).value = str(sheet_in2.cell(row = i+2, column = 19).value)
			sheet_out.cell(row = serial_number+1, column = 22).value = str(sheet_in2.cell(row = i+2, column = 20).value)
			sheet_out.cell(row = serial_number+1, column = 23).value = str(sheet_in2.cell(row = i+2, column = 21).value)
			sheet_out.cell(row = serial_number+1, column = 24).value = str(sheet_in2.cell(row = i+2, column = 22).value)
			sheet_out.cell(row = serial_number+1, column = 25).value = str(sheet_in2.cell(row = i+2, column = 24).value)
			sheet_out.cell(row = serial_number+1, column = 26).value = str(sheet_in2.cell(row = i+2, column = 25).value)
			sheet_out.cell(row = serial_number+1, column = 27).value = str(sheet_in2.cell(row = i+2, column = 31).value)
			sheet_out.cell(row = serial_number+1, column = 28).value = str(sheet_in2.cell(row = i+2, column = 26).value)
			sheet_out.cell(row = serial_number+1, column = 29).value = str(sheet_in2.cell(row = i+2, column = 27).value)
			sheet_out.cell(row = serial_number+1, column = 30).value = str(sheet_in2.cell(row = i+2, column = 28).value)
			sheet_out.cell(row = serial_number+1, column = 31).value = str(sheet_in2.cell(row = i+2, column = 29).value)
			sheet_out.cell(row = serial_number+1, column = 32).value = str(sheet_in2.cell(row = i+2, column = 30).value)
			sheet_out.cell(row = serial_number+1, column = 33).value = str(sheet_in2.cell(row = i+2, column = 32).value)
			sheet_out.cell(row = serial_number+1, column = 34).value = str(sheet_in2.cell(row = i+2, column = 33).value)
			sheet_out.cell(row = serial_number+1, column = 35).value = str(sheet_in2.cell(row = i+2, column = 39).value)
			sheet_out.cell(row = serial_number+1, column = 36).value = str(sheet_in2.cell(row = i+2, column = 34).value)
			sheet_out.cell(row = serial_number+1, column = 37).value = str(sheet_in2.cell(row = i+2, column = 35).value)
			sheet_out.cell(row = serial_number+1, column = 38).value = str(sheet_in2.cell(row = i+2, column = 36).value)
			sheet_out.cell(row = serial_number+1, column = 39).value = str(sheet_in2.cell(row = i+2, column = 37).value)
			sheet_out.cell(row = serial_number+1, column = 40).value = str(sheet_in2.cell(row = i+2, column = 38).value)
			sheet_out.cell(row = serial_number+1, column = 41).value = str(sheet_in2.cell(row = i+2, column = 40).value)
			sheet_out.cell(row = serial_number+1, column = 42).value = str(sheet_in2.cell(row = i+2, column = 41).value)
			sheet_out.cell(row = serial_number+1, column = 43).value = str(sheet_in2.cell(row = i+2, column = 47).value)
			sheet_out.cell(row = serial_number+1, column = 44).value = str(sheet_in2.cell(row = i+2, column = 42).value)
			sheet_out.cell(row = serial_number+1, column = 45).value = str(sheet_in2.cell(row = i+2, column = 43).value)
			sheet_out.cell(row = serial_number+1, column = 46).value = str(sheet_in2.cell(row = i+2, column = 44).value)
			sheet_out.cell(row = serial_number+1, column = 47).value = str(sheet_in2.cell(row = i+2, column = 45).value)
			sheet_out.cell(row = serial_number+1, column = 48).value = str(sheet_in2.cell(row = i+2, column = 46).value)
			sheet_out.cell(row = serial_number+1, column = 49).value = str(sheet_in2.cell(row = i+2, column = 48).value)
			sheet_out.cell(row = serial_number+1, column = 50).value = str(sheet_in2.cell(row = i+2, column = 49).value)
			sheet_out.cell(row = serial_number+1, column = 51).value = str(sheet_in2.cell(row = i+2, column = 55).value)
			sheet_out.cell(row = serial_number+1, column = 52).value = str(sheet_in2.cell(row = i+2, column = 50).value)
			sheet_out.cell(row = serial_number+1, column = 53).value = str(sheet_in2.cell(row = i+2, column = 51).value)
			sheet_out.cell(row = serial_number+1, column = 54).value = str(sheet_in2.cell(row = i+2, column = 52).value)
			sheet_out.cell(row = serial_number+1, column = 55).value = str(sheet_in2.cell(row = i+2, column = 53).value)
			sheet_out.cell(row = serial_number+1, column = 56).value = str(sheet_in2.cell(row = i+2, column = 54).value)
			sheet_out.cell(row = serial_number+1, column = 57).value = "NPH"
			sheet_out.cell(row = serial_number+1, column = 58).value = "SPH"
	else:
		pass
	if(num_rows_sn>0):
		for i in range(num_rows_sn):
			serial_number = int(sheet_in3.cell(row = i+2, column = 1).value)
			
			sheet_out.cell(row = serial_number+1, column = 1).value = serial_number
			sheet_out.cell(row = serial_number+1, column = 2).value = str(sheet_in3.cell(row = i+2, column = 2).value)
			sheet_out.cell(row = serial_number+1, column = 3).value = str(sheet_in3.cell(row = i+2, column = 3).value)
			sheet_out.cell(row = serial_number+1, column = 4).value = str(sheet_in3.cell(row = i+2, column = 4).value)
			sheet_out.cell(row = serial_number+1, column = 5).value = str(sheet_in3.cell(row = i+2, column = 5).value)
			sheet_out.cell(row = serial_number+1, column = 6).value = str(sheet_in3.cell(row = i+2, column = 6).value)
			sheet_out.cell(row = serial_number+1, column = 6).fill = PatternFill(fgColor='FFEE08', fill_type = 'solid')
			sheet_out.cell(row = serial_number+1, column = 7).value = '-'
			a=float(sheet_in3.cell(row = i+2, column = 7).value)
			sheet_out.cell(row = serial_number+1, column = 8).value = format(a, '.2f')
			sheet_out.cell(row = serial_number+1, column = 8).fill = PatternFill(fgColor='FFEE08', fill_type = 'solid')
			sheet_out.cell(row = serial_number+1, column = 8).font = Font(bold=True)
			sheet_out.cell(row = serial_number+1, column = 9).value = str(sheet_in3.cell(row = i+2, column = 8).value)
			sheet_out.cell(row = serial_number+1, column = 10).value = str(sheet_in3.cell(row = i+2, column = 9).value)
			sheet_out.cell(row = serial_number+1, column = 11).value = str(sheet_in3.cell(row = i+2, column = 15).value)
			sheet_out.cell(row = serial_number+1, column = 12).value = str(sheet_in3.cell(row = i+2, column = 10).value)
			sheet_out.cell(row = serial_number+1, column = 13).value = str(sheet_in3.cell(row = i+2, column = 11).value)
			sheet_out.cell(row = serial_number+1, column = 14).value = str(sheet_in3.cell(row = i+2, column = 12).value)
			sheet_out.cell(row = serial_number+1, column = 15).value = str(sheet_in3.cell(row = i+2, column = 13).value)
			sheet_out.cell(row = serial_number+1, column = 16).value = str(sheet_in3.cell(row = i+2, column = 14).value)
			sheet_out.cell(row = serial_number+1, column = 17).value = str(sheet_in3.cell(row = i+2, column = 16).value)
			sheet_out.cell(row = serial_number+1, column = 18).value = str(sheet_in3.cell(row = i+2, column = 17).value)
			sheet_out.cell(row = serial_number+1, column = 19).value = str(sheet_in3.cell(row = i+2, column = 23).value)
			sheet_out.cell(row = serial_number+1, column = 20).value = str(sheet_in3.cell(row = i+2, column = 18).value)
			sheet_out.cell(row = serial_number+1, column = 21).value = str(sheet_in3.cell(row = i+2, column = 19).value)
			sheet_out.cell(row = serial_number+1, column = 22).value = str(sheet_in3.cell(row = i+2, column = 20).value)
			sheet_out.cell(row = serial_number+1, column = 23).value = str(sheet_in3.cell(row = i+2, column = 21).value)
			sheet_out.cell(row = serial_number+1, column = 24).value = str(sheet_in3.cell(row = i+2, column = 22).value)
			sheet_out.cell(row = serial_number+1, column = 25).value = str(sheet_in3.cell(row = i+2, column = 24).value)
			sheet_out.cell(row = serial_number+1, column = 26).value = str(sheet_in3.cell(row = i+2, column = 25).value)
			sheet_out.cell(row = serial_number+1, column = 27).value = str(sheet_in3.cell(row = i+2, column = 31).value)
			sheet_out.cell(row = serial_number+1, column = 28).value = str(sheet_in3.cell(row = i+2, column = 26).value)
			sheet_out.cell(row = serial_number+1, column = 29).value = str(sheet_in3.cell(row = i+2, column = 27).value)
			sheet_out.cell(row = serial_number+1, column = 30).value = str(sheet_in3.cell(row = i+2, column = 28).value)
			sheet_out.cell(row = serial_number+1, column = 31).value = str(sheet_in3.cell(row = i+2, column = 29).value)
			sheet_out.cell(row = serial_number+1, column = 32).value = str(sheet_in3.cell(row = i+2, column = 30).value)
			sheet_out.cell(row = serial_number+1, column = 33).value = str(sheet_in3.cell(row = i+2, column = 32).value)
			sheet_out.cell(row = serial_number+1, column = 34).value = str(sheet_in3.cell(row = i+2, column = 33).value)
			sheet_out.cell(row = serial_number+1, column = 35).value = str(sheet_in3.cell(row = i+2, column = 39).value)
			sheet_out.cell(row = serial_number+1, column = 36).value = str(sheet_in3.cell(row = i+2, column = 34).value)
			sheet_out.cell(row = serial_number+1, column = 37).value = str(sheet_in3.cell(row = i+2, column = 35).value)
			sheet_out.cell(row = serial_number+1, column = 38).value = str(sheet_in3.cell(row = i+2, column = 36).value)
			sheet_out.cell(row = serial_number+1, column = 39).value = str(sheet_in3.cell(row = i+2, column = 37).value)
			sheet_out.cell(row = serial_number+1, column = 40).value = str(sheet_in3.cell(row = i+2, column = 38).value)
			sheet_out.cell(row = serial_number+1, column = 41).value = str(sheet_in3.cell(row = i+2, column = 40).value)
			sheet_out.cell(row = serial_number+1, column = 42).value = str(sheet_in3.cell(row = i+2, column = 41).value)
			sheet_out.cell(row = serial_number+1, column = 43).value = str(sheet_in3.cell(row = i+2, column = 47).value)
			sheet_out.cell(row = serial_number+1, column = 44).value = str(sheet_in3.cell(row = i+2, column = 42).value)
			sheet_out.cell(row = serial_number+1, column = 45).value = str(sheet_in3.cell(row = i+2, column = 43).value)
			sheet_out.cell(row = serial_number+1, column = 46).value = str(sheet_in3.cell(row = i+2, column = 44).value)
			sheet_out.cell(row = serial_number+1, column = 47).value = str(sheet_in3.cell(row = i+2, column = 45).value)
			sheet_out.cell(row = serial_number+1, column = 48).value = str(sheet_in3.cell(row = i+2, column = 46).value)
			sheet_out.cell(row = serial_number+1, column = 49).value = str(sheet_in3.cell(row = i+2, column = 48).value)
			sheet_out.cell(row = serial_number+1, column = 50).value = str(sheet_in3.cell(row = i+2, column = 49).value)
			sheet_out.cell(row = serial_number+1, column = 51).value = str(sheet_in3.cell(row = i+2, column = 55).value)
			sheet_out.cell(row = serial_number+1, column = 52).value = str(sheet_in3.cell(row = i+2, column = 50).value)
			sheet_out.cell(row = serial_number+1, column = 53).value = str(sheet_in3.cell(row = i+2, column = 51).value)
			sheet_out.cell(row = serial_number+1, column = 54).value = str(sheet_in3.cell(row = i+2, column = 52).value)
			sheet_out.cell(row = serial_number+1, column = 55).value = str(sheet_in3.cell(row = i+2, column = 53).value)
			sheet_out.cell(row = serial_number+1, column = 56).value = str(sheet_in3.cell(row = i+2, column = 54).value)
			sheet_out.cell(row = serial_number+1, column = 57).value = "SPH"
			sheet_out.cell(row = serial_number+1, column = 58).value = "NPH"
	else:
		pass
	if(num_rows_ss>0):
		for i in range(num_rows_ss):
			serial_number = int(sheet_in4.cell(row = i+2, column = 1).value)
			sheet_out.cell(row = serial_number+1, column = 1).value = serial_number
			sheet_out.cell(row = serial_number+1, column = 2).value = str(sheet_in4.cell(row = i+2, column = 2).value)
			sheet_out.cell(row = serial_number+1, column = 3).value = str(sheet_in4.cell(row = i+2, column = 3).value)
			sheet_out.cell(row = serial_number+1, column = 4).value = str(sheet_in4.cell(row = i+2, column = 4).value)
			sheet_out.cell(row = serial_number+1, column = 5).value = str(sheet_in4.cell(row = i+2, column = 5).value)
			sheet_out.cell(row = serial_number+1, column = 6).value = str(sheet_in4.cell(row = i+2, column = 6).value)
			sheet_out.cell(row = serial_number+1, column = 6).fill = PatternFill(fgColor='FFEE08', fill_type = 'solid')
			sheet_out.cell(row = serial_number+1, column = 7).value = str(sheet_in4.cell(row = i+2, column = 7).value)
			sheet_out.cell(row = serial_number+1, column = 7).fill = PatternFill(fgColor='FFEE08', fill_type = 'solid')
			a=float(str(sheet_in4.cell(row = i+2, column = 8).value))
			sheet_out.cell(row = serial_number+1, column = 8).value = format(a, '.2f')
			sheet_out.cell(row = serial_number+1, column = 8).fill = PatternFill(fgColor='FFEE08', fill_type = 'solid')
			sheet_out.cell(row = serial_number+1, column = 8).font = Font(bold=True)
			sheet_out.cell(row = serial_number+1, column = 9).value = str(sheet_in4.cell(row = i+2, column = 9).value)
			sheet_out.cell(row = serial_number+1, column = 10).value = str(sheet_in4.cell(row = i+2, column = 10).value)
			sheet_out.cell(row = serial_number+1, column = 11).value = str(sheet_in4.cell(row = i+2, column = 16).value)
			sheet_out.cell(row = serial_number+1, column = 12).value = str(sheet_in4.cell(row = i+2, column = 11).value)
			sheet_out.cell(row = serial_number+1, column = 13).value = str(sheet_in4.cell(row = i+2, column = 12).value)
			sheet_out.cell(row = serial_number+1, column = 14).value = str(sheet_in4.cell(row = i+2, column = 13).value)
			sheet_out.cell(row = serial_number+1, column = 15).value = str(sheet_in4.cell(row = i+2, column = 14).value)
			sheet_out.cell(row = serial_number+1, column = 16).value = str(sheet_in4.cell(row = i+2, column = 15).value)
			sheet_out.cell(row = serial_number+1, column = 17).value = str(sheet_in4.cell(row = i+2, column = 17).value)
			sheet_out.cell(row = serial_number+1, column = 18).value = str(sheet_in4.cell(row = i+2, column = 18).value)
			sheet_out.cell(row = serial_number+1, column = 19).value = str(sheet_in4.cell(row = i+2, column = 24).value)
			sheet_out.cell(row = serial_number+1, column = 20).value = str(sheet_in4.cell(row = i+2, column = 19).value)
			sheet_out.cell(row = serial_number+1, column = 21).value = str(sheet_in4.cell(row = i+2, column = 20).value)
			sheet_out.cell(row = serial_number+1, column = 22).value = str(sheet_in4.cell(row = i+2, column = 21).value)
			sheet_out.cell(row = serial_number+1, column = 23).value = str(sheet_in4.cell(row = i+2, column = 22).value)
			sheet_out.cell(row = serial_number+1, column = 24).value = str(sheet_in4.cell(row = i+2, column = 23).value)
			sheet_out.cell(row = serial_number+1, column = 25).value = str(sheet_in4.cell(row = i+2, column = 25).value)
			sheet_out.cell(row = serial_number+1, column = 26).value = str(sheet_in4.cell(row = i+2, column = 26).value)
			sheet_out.cell(row = serial_number+1, column = 27).value = str(sheet_in4.cell(row = i+2, column = 32).value)
			sheet_out.cell(row = serial_number+1, column = 28).value = str(sheet_in4.cell(row = i+2, column = 27).value)
			sheet_out.cell(row = serial_number+1, column = 29).value = str(sheet_in4.cell(row = i+2, column = 28).value)
			sheet_out.cell(row = serial_number+1, column = 30).value = str(sheet_in4.cell(row = i+2, column = 29).value)
			sheet_out.cell(row = serial_number+1, column = 31).value = str(sheet_in4.cell(row = i+2, column = 30).value)
			sheet_out.cell(row = serial_number+1, column = 32).value = str(sheet_in4.cell(row = i+2, column = 31).value)
			sheet_out.cell(row = serial_number+1, column = 33).value = str(sheet_in4.cell(row = i+2, column = 33).value)
			sheet_out.cell(row = serial_number+1, column = 34).value = str(sheet_in4.cell(row = i+2, column = 34).value)
			sheet_out.cell(row = serial_number+1, column = 35).value = str(sheet_in4.cell(row = i+2, column = 40).value)
			sheet_out.cell(row = serial_number+1, column = 36).value = str(sheet_in4.cell(row = i+2, column = 35).value)
			sheet_out.cell(row = serial_number+1, column = 37).value = str(sheet_in4.cell(row = i+2, column = 36).value)
			sheet_out.cell(row = serial_number+1, column = 38).value = str(sheet_in4.cell(row = i+2, column = 37).value)
			sheet_out.cell(row = serial_number+1, column = 39).value = str(sheet_in4.cell(row = i+2, column = 38).value)
			sheet_out.cell(row = serial_number+1, column = 40).value = str(sheet_in4.cell(row = i+2, column = 39).value)
			sheet_out.cell(row = serial_number+1, column = 41).value = str(sheet_in4.cell(row = i+2, column = 41).value)
			sheet_out.cell(row = serial_number+1, column = 42).value = str(sheet_in4.cell(row = i+2, column = 42).value)
			sheet_out.cell(row = serial_number+1, column = 43).value = str(sheet_in4.cell(row = i+2, column = 48).value)
			sheet_out.cell(row = serial_number+1, column = 44).value = str(sheet_in4.cell(row = i+2, column = 43).value)
			sheet_out.cell(row = serial_number+1, column = 45).value = str(sheet_in4.cell(row = i+2, column = 44).value)
			sheet_out.cell(row = serial_number+1, column = 46).value = str(sheet_in4.cell(row = i+2, column = 45).value)
			sheet_out.cell(row = serial_number+1, column = 47).value = str(sheet_in4.cell(row = i+2, column = 46).value)
			sheet_out.cell(row = serial_number+1, column = 48).value = str(sheet_in4.cell(row = i+2, column = 47).value)
			sheet_out.cell(row = serial_number+1, column = 49).value = str(sheet_in4.cell(row = i+2, column = 49).value)
			sheet_out.cell(row = serial_number+1, column = 50).value = str(sheet_in4.cell(row = i+2, column = 50).value)
			sheet_out.cell(row = serial_number+1, column = 51).value = str(sheet_in4.cell(row = i+2, column = 56).value)
			sheet_out.cell(row = serial_number+1, column = 52).value = str(sheet_in4.cell(row = i+2, column = 51).value)
			sheet_out.cell(row = serial_number+1, column = 53).value = str(sheet_in4.cell(row = i+2, column = 52).value)
			sheet_out.cell(row = serial_number+1, column = 54).value = str(sheet_in4.cell(row = i+2, column = 53).value)
			sheet_out.cell(row = serial_number+1, column = 55).value = str(sheet_in4.cell(row = i+2, column = 54).value)
			sheet_out.cell(row = serial_number+1, column = 56).value = str(sheet_in4.cell(row = i+2, column = 55).value)
			sheet_out.cell(row = serial_number+1, column = 57).value = "SPH"
			sheet_out.cell(row = serial_number+1, column = 58).value = "SPH"
	else:
		pass
	num_rows = num_rows_nn+num_rows_ns+num_rows_sn+num_rows_ss
	print(num_rows_nn,num_rows_ns,num_rows_sn,num_rows_ss)


	print(num_rows)
	for j in range(num_rows):
		for k in range(58):
			mode = str(sheet_out.cell(row = j+2, column = k+1).value)
			if(mode=='Air'):
				sheet_out.cell(row = j+2, column = k+1).fill = PatternFill(fgColor='000000FF', fill_type = 'solid')
			if(mode=='None'):
				sheet_out.cell(row = j+2, column = k+1).value = ""

	c = sheet_out['I2']
	sheet_out.freeze_panes = c

	fp_out.save('C:\\web2py_win\\web2py\\applications\\IndiaPost_UI\\Source Code\\IndiaPost\\consolidate.xlsx')

	'''
	workbook = xlsxwriter.Workbook(curDir+'Consolidate_Air.xlsx')
	worksheet = workbook.add_worksheet('Sheet')
	worksheet.freeze_panes(1, 0) # # Freeze the first row.
	worksheet.freeze_panes(0, 1) # # Freeze the first column.
	worksheet.freeze_panes(0, 2)
	worksheet.freeze_panes(0, 3)
	worksheet.freeze_panes(0, 4)
	worksheet.freeze_panes(0, 5)
	worksheet.freeze_panes(0, 6)
	worksheet.freeze_panes(0, 7)
	worksheet.freeze_panes(0, 8)
	workbook.close()
	'''

#fp_out.save('C:\\web2py_win\\web2py\\applications\\IndiaPost\\private\\Output.xlsx')
	

if __name__ == '__main__':
	main()
		